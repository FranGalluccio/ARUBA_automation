"""
Genera il file Excel con la documentazione completa della suite di test PEC Aruba.
"""
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_TEXT
from openpyxl.worksheet.table import Table, TableStyleInfo
import os

OUTPUT_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "suite_test_pec.xlsx")

# ── PALETTE ───────────────────────────────────────────────────────────────────
C_DARK_BLUE   = "1F3864"   # header principale
C_MID_BLUE    = "2E75B6"   # header di sezione
C_LIGHT_BLUE  = "D6E4F0"   # riga pari modulo Login/Impostazioni
C_GREEN       = "375623"   # header modulo Messaggi
C_LIGHT_GREEN = "E2EFDA"
C_PURPLE      = "4B0082"   # header modulo Calendario
C_LIGHT_PUR   = "EAE0F5"
C_ORANGE      = "843C0C"   # header modulo Contatti
C_LIGHT_ORA   = "FCE4D6"
C_TEAL        = "1D5F5F"   # header modulo Impostazioni
C_LIGHT_TEA   = "D9EEE8"
C_MAROON      = "7B2D00"   # header modulo Archivio
C_LIGHT_MAR   = "F9E8E0"
C_INDIGO      = "283593"   # header modulo Conservazione
C_LIGHT_IND   = "E8EAF6"
C_OLIVE       = "4E6C0A"   # header modulo Fatture
C_LIGHT_OLI   = "F1F8E9"
C_PLUM        = "6A1B4D"   # header modulo Accessi
C_LIGHT_PLU   = "F3E5F5"
C_WHITE       = "FFFFFF"
C_LIGHT_GRAY  = "F2F2F2"
C_GOLD        = "FFD700"

THIN  = Side(style="thin",   color="BFBFBF")
MED   = Side(style="medium", color="595959")
BORDER_THIN = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BORDER_MED_TOP = Border(left=THIN, right=THIN, top=MED, bottom=THIN)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color=C_DARK_BLUE, size=10, italic=False):
    return Font(name="Calibri", bold=bold, color=color, size=size, italic=italic)

def align(h="left", v="top", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


# ── TEST DATA ─────────────────────────────────────────────────────────────────
# Columns: (id, modulo, file, funzione, descrizione, passi, risultato_atteso)

TESTS = [

    # ══════════════════════════════════════════════════════════════════════════
    # LOGIN
    # ══════════════════════════════════════════════════════════════════════════
    (1, "Login", "test_login_01.py", "test_logout",
     "Verifica che il logout dalla casella PEC funzioni correttamente.",
     "1. Accedere con credenziali valide.\n"
     "2. Attendere caricamento della webmail.\n"
     "3. Cliccare il pulsante dell'account (username).\n"
     "4. Cliccare il pulsante 'Esci' dal menu dropdown.\n"
     "5. Attendere il reindirizzamento alla pagina di login.\n"
     "6. Scattare screenshot della pagina di login.\n"
     "7. Verificare che il campo username sia visibile (assert).",
     "L'URL non contiene più 'INBOX'. Il campo di input username è visibile nella pagina di login."),

    (2, "Login", "test_login_02.py", "test_login_credenziali_errate",
     "Verifica che il login con credenziali non valide mostri un messaggio di errore.",
     "1. Aprire la pagina di login della webmail PEC.\n"
     "2. Accettare il banner cookie se presente.\n"
     "3. Inserire un username inesistente nel campo email.\n"
     "4. Inserire una password errata nel campo password.\n"
     "5. Cliccare il pulsante Login.\n"
     "6. Attendere 2 secondi.\n"
     "7. Verificare che l'URL non contenga 'INBOX'.\n"
     "8. Scattare screenshot.\n"
     "9. Verificare la visibilità del messaggio di errore (assert).",
     "L'URL rimane sulla pagina di login (no 'INBOX'). Il div 'errorWeb' è visibile e contiene testo di errore relativo alle credenziali errate."),

    # ══════════════════════════════════════════════════════════════════════════
    # MESSAGGI
    # ══════════════════════════════════════════════════════════════════════════
    (3, "Messaggi", "test_messaggi_01.py", "test_messaggio_con_allegato",
     "Verifica l'invio di un messaggio PEC con allegato e la sua corretta visualizzazione.",
     "1. Accedere con credenziali valide.\n"
     "2. Cliccare 'Nuovo messaggio'.\n"
     "3. Inserire l'indirizzo del destinatario principale.\n"
     "4. Inserire oggetto e corpo del messaggio.\n"
     "5. Allegare un file PDF tramite il pulsante allegati.\n"
     "6. Inviare il messaggio.\n"
     "7. Attendere la ricezione nella cartella In arrivo.\n"
     "8. Aprire il messaggio ricevuto.\n"
     "9. Scattare screenshot.\n"
     "10. Verificare la presenza dell'allegato nel messaggio (assert).",
     "Il messaggio è presente in In arrivo. L'allegato è visibile all'interno del messaggio ricevuto."),

    (4, "Messaggi", "test_messaggi_02.py", "test_messaggio_inoltrato",
     "Verifica l'inoltro di un messaggio PEC ricevuto.",
     "1. Accedere con credenziali valide.\n"
     "2. Inviare un messaggio a se stessi con oggetto univoco.\n"
     "3. Attendere la ricezione del messaggio.\n"
     "4. Aprire il messaggio ricevuto.\n"
     "5. Cliccare il pulsante 'Inoltra'.\n"
     "6. Inserire il destinatario dell'inoltro.\n"
     "7. Inviare il messaggio inoltrato.\n"
     "8. Scattare screenshot.\n"
     "9. Verificare la comparsa del toast di conferma (assert).",
     "Il messaggio inoltrato è inviato correttamente. Il toast di conferma è visibile. L'oggetto contiene il prefisso di inoltro ('Fwd:', 'I:', ecc.)."),

    (5, "Messaggi", "test_messaggi_03.py", "test_messaggio_in_bozza",
     "Verifica il salvataggio di un messaggio come bozza e il successivo invio.",
     "1. Accedere con credenziali valide.\n"
     "2. Cliccare 'Nuovo messaggio'.\n"
     "3. Inserire destinatario, oggetto e corpo.\n"
     "4. Cliccare 'Salva bozza' (o chiudere la finestra senza inviare).\n"
     "5. Navigare nella cartella Bozze.\n"
     "6. Verificare che la bozza sia presente.\n"
     "7. Aprire la bozza e inviarla.\n"
     "8. Scattare screenshot.",
     "La bozza è visibile nella cartella Bozze. Dopo l'invio, la bozza non è più presente nella cartella Bozze."),

    (6, "Messaggi", "test_messaggi_04.py", "test_messaggio_importato",
     "Verifica l'importazione di messaggi da file nella casella PEC.",
     "1. Accedere con credenziali valide.\n"
     "2. Aprire il menu di importazione messaggi.\n"
     "3. Selezionare il file da importare.\n"
     "4. Avviare il processo di importazione.\n"
     "5. Attendere il completamento dell'importazione.\n"
     "6. Scattare screenshot.\n"
     "7. Verificare il messaggio di conferma (assert).",
     "Il processo di importazione si completa senza errori. Viene mostrato un messaggio di successo o conferma."),

    (7, "Messaggi", "test_messaggi_05.py", "test_cartella",
     "Verifica la creazione, la modifica e l'eliminazione di una cartella personalizzata.",
     "1. Accedere con credenziali valide.\n"
     "2. Creare una nuova cartella con nome univoco.\n"
     "3. Verificare il toast di conferma della creazione.\n"
     "4. Rinominare la cartella con un nuovo nome.\n"
     "5. Verificare il toast di conferma della modifica.\n"
     "6. Eliminare la cartella.\n"
     "7. Scattare screenshot.\n"
     "8. Verificare il toast di conferma dell'eliminazione (assert).",
     "La cartella è creata, rinominata ed eliminata con successo. Per ogni operazione compare un toast di conferma."),

    (8, "Messaggi", "test_messaggi_06.py", "test_ripristino_messaggi",
     "Verifica il ripristino di messaggi dal cestino alla cartella In arrivo.",
     "1. Accedere con credenziali valide.\n"
     "2. Spostare un messaggio nel cestino.\n"
     "3. Navigare nella cartella Cestino.\n"
     "4. Selezionare il/i messaggio/i da ripristinare.\n"
     "5. Cliccare 'Ripristina' o 'Sposta in In arrivo'.\n"
     "6. Verificare che i messaggi non siano più nel cestino.\n"
     "7. Scattare screenshot.\n"
     "8. Navigare in In arrivo e verificare la presenza dei messaggi (assert).",
     "I messaggi ripristinati non sono più nel cestino e sono visibili nella cartella In arrivo."),

    (9, "Messaggi", "test_messaggi_07.py", "test_messaggi_preferiti_pinnati",
     "Verifica l'aggiunta di messaggi ai preferiti e la funzione 'Pinna messaggio'.",
     "1. Accedere con credenziali valide.\n"
     "2. Selezionare un messaggio dall'In arrivo.\n"
     "3. Aggiungere il messaggio ai preferiti (stella).\n"
     "4. Verificare che il messaggio compaia nella cartella Preferiti/Speciali.\n"
     "5. Selezionare un messaggio e cliccare 'Pinna messaggio'.\n"
     "6. Scattare screenshot.\n"
     "7. Verificare la presenza del messaggio pinnato (assert).",
     "Il messaggio è aggiunto ai preferiti e compare nella sezione Preferiti. Il messaggio pinnato è visibile in cima alla lista."),

    (10, "Messaggi", "test_messaggi_08.py", "test_etichetta",
     "Verifica la creazione di un'etichetta, la sua applicazione ai messaggi e l'eliminazione.",
     "1. Accedere con credenziali valide.\n"
     "2. Aprire le impostazioni etichette.\n"
     "3. Creare una nuova etichetta con nome univoco e colore.\n"
     "4. Tornare in In arrivo e selezionare un messaggio.\n"
     "5. Applicare l'etichetta appena creata al messaggio.\n"
     "6. Verificare che l'etichetta sia visibile sul messaggio.\n"
     "7. Scattare screenshot.\n"
     "8. Eliminare l'etichetta (cleanup).",
     "L'etichetta è creata e applicata con successo al messaggio. Dopo l'eliminazione dell'etichetta, essa non è più visibile."),

    (11, "Messaggi", "test_messaggi_09.py", "test_segna_come_letto_da_leggere",
     "Verifica le funzioni 'Segna come letto' e 'Segna come da leggere' sui messaggi.",
     "1. Accedere con credenziali valide.\n"
     "2. Selezionare un messaggio non letto dall'In arrivo.\n"
     "3. Cliccare 'Segna come letto'.\n"
     "4. Verificare che il messaggio non sia più in grassetto (letto).\n"
     "5. Cliccare 'Segna come da leggere' sullo stesso messaggio.\n"
     "6. Scattare screenshot.\n"
     "7. Verificare che il messaggio torni in grassetto (non letto, assert).",
     "I pulsanti 'Segna come letto' e 'Segna come da leggere' sono visibili e funzionanti. Lo stato del messaggio cambia correttamente."),

    (12, "Messaggi", "test_messaggi_10.py", "test_messaggio_alta_priorita",
     "Verifica l'invio e la ricezione di un messaggio con priorità alta.",
     "1. Accedere con credenziali valide.\n"
     "2. Cliccare 'Nuovo messaggio'.\n"
     "3. Inserire destinatario, oggetto e corpo.\n"
     "4. Impostare la priorità del messaggio su 'Alta'.\n"
     "5. Inviare il messaggio.\n"
     "6. Attendere la ricezione in In arrivo.\n"
     "7. Aprire il messaggio ricevuto.\n"
     "8. Scattare screenshot.\n"
     "9. Verificare la presenza dell'indicatore di alta priorità (assert).",
     "Il messaggio è ricevuto con l'indicatore di alta priorità visibile (simbolo '!' o equivalente)."),

    (13, "Messaggi", "test_messaggi_11.py", "test_risposta_messaggio",
     "Verifica la risposta a un messaggio PEC ricevuto.",
     "1. Accedere con credenziali valide.\n"
     "2. Inviare un messaggio a se stessi.\n"
     "3. Attendere la ricezione del messaggio.\n"
     "4. Aprire il messaggio ricevuto.\n"
     "5. Cliccare il pulsante 'Rispondi'.\n"
     "6. Verificare che la finestra di risposta sia aperta con 'Re:' nell'oggetto.\n"
     "7. Inviare la risposta.\n"
     "8. Scattare screenshot.",
     "La finestra di risposta si apre correttamente. L'oggetto contiene il prefisso 'Re:'. La risposta viene inviata con successo."),

    (14, "Messaggi", "test_messaggi_12.py", "test_invio_messaggio_semplice",
     "Verifica l'invio di un messaggio PEC semplice senza allegati.",
     "1. Accedere con credenziali valide.\n"
     "2. Cliccare 'Nuovo messaggio'.\n"
     "3. Inserire l'indirizzo del destinatario.\n"
     "4. Inserire oggetto e corpo del messaggio.\n"
     "5. Cliccare 'Invia'.\n"
     "6. Scattare screenshot.\n"
     "7. Verificare la comparsa del toast di conferma invio (assert).",
     "Il messaggio è inviato con successo. Appare un toast di conferma dell'invio."),

    (15, "Messaggi", "test_messaggi_13.py", "test_risposta_a_tutti",
     "Verifica la funzione 'Rispondi a tutti' su un messaggio ricevuto.",
     "1. Accedere con credenziali valide.\n"
     "2. Inviare un messaggio a se stessi.\n"
     "3. Attendere la ricezione del messaggio.\n"
     "4. Aprire il messaggio ricevuto.\n"
     "5. Cliccare il pulsante 'Rispondi a tutti'.\n"
     "6. Verificare che la finestra di risposta sia aperta con 'Re:' nell'oggetto.\n"
     "7. Verificare che tutti i destinatari originali siano inclusi.\n"
     "8. Inviare la risposta a tutti.\n"
     "9. Scattare screenshot.",
     "La finestra di risposta si apre con l'oggetto 'Re: ...' e tutti i destinatari originali inclusi. La risposta viene inviata con successo."),

    (16, "Messaggi", "test_messaggi_14.py", "test_elimina_messaggio",
     "Verifica l'eliminazione di un messaggio e la sua presenza nel cestino.",
     "1. Accedere con credenziali valide.\n"
     "2. Inviare un messaggio a se stessi con oggetto univoco.\n"
     "3. Attendere la ricezione del messaggio.\n"
     "4. Selezionare il messaggio e cliccare 'Elimina'.\n"
     "5. Navigare nella cartella Cestino.\n"
     "6. Scattare screenshot.\n"
     "7. Verificare che il messaggio eliminato sia presente nel cestino (assert).",
     "Il messaggio eliminato non è più in In arrivo ma è visibile nella cartella Cestino."),

    (17, "Messaggi", "test_messaggi_15.py", "test_sposta_messaggio_in_cartella",
     "Verifica lo spostamento di un messaggio in una cartella personalizzata.",
     "1. Accedere con credenziali valide.\n"
     "2. Creare una nuova cartella con nome univoco.\n"
     "3. Inviare un messaggio a se stessi.\n"
     "4. Attendere la ricezione del messaggio.\n"
     "5. Selezionare il messaggio e cliccare 'Sposta in...'.\n"
     "6. Selezionare la cartella di destinazione.\n"
     "7. Verificare che il messaggio non sia più in In arrivo.\n"
     "8. Scattare screenshot.\n"
     "9. Navigare nella cartella di destinazione e verificare la presenza del messaggio (assert).",
     "Il messaggio è spostato correttamente nella cartella di destinazione e non è più visibile in In arrivo."),

    (18, "Messaggi", "test_messaggi_16.py", "test_copia_messaggio_in_cartella",
     "Verifica la copia di un messaggio in una cartella personalizzata.",
     "1. Accedere con credenziali valide.\n"
     "2. Creare una nuova cartella con nome univoco.\n"
     "3. Aprire un messaggio esistente in In arrivo.\n"
     "4. Selezionare 'Copia in...' dal menu azioni.\n"
     "5. Selezionare la cartella di destinazione.\n"
     "6. Verificare che il messaggio originale sia ancora in In arrivo.\n"
     "7. Scattare screenshot.\n"
     "8. Navigare nella cartella di destinazione e verificare la copia (assert).",
     "Il messaggio è copiato nella cartella di destinazione. Il messaggio originale rimane in In arrivo."),

    (19, "Messaggi", "test_messaggi_17.py", "test_ricerca_messaggi",
     "Verifica la funzione di ricerca messaggi nella casella PEC.",
     "1. Accedere con credenziali valide.\n"
     "2. Inviare un messaggio a se stessi con oggetto univoco.\n"
     "3. Attendere la ricezione del messaggio.\n"
     "4. Inserire il testo univoco nella barra di ricerca.\n"
     "5. Avviare la ricerca.\n"
     "6. Scattare screenshot.\n"
     "7. Verificare che il messaggio cercato appaia nei risultati (assert).",
     "Il messaggio con l'oggetto univoco è trovato e visibile nei risultati di ricerca."),

    (20, "Messaggi", "test_messaggi_18.py", "test_filtri_inbox",
     "Verifica la presenza e il funzionamento del componente filtri messaggi.",
     "1. Accedere con credenziali valide.\n"
     "2. Navigare nella cartella In arrivo.\n"
     "3. Individuare il pulsante/dropdown filtri.\n"
     "4. Cliccare sul filtro per aprire le opzioni.\n"
     "5. Selezionare un'opzione di filtro (es. 'Non letti').\n"
     "6. Scattare screenshot.\n"
     "7. Verificare che la lista messaggi sia filtrata correttamente (assert).",
     "Il componente filtri è visibile. Dopo la selezione di un filtro, la lista messaggi mostra solo i messaggi corrispondenti."),

    (21, "Messaggi", "test_messaggi_19.py", "test_invio_messaggio_con_cc",
     "Verifica l'invio di un messaggio PEC con destinatario in copia (CC).",
     "1. Accedere con credenziali valide.\n"
     "2. Cliccare 'Nuovo messaggio'.\n"
     "3. Inserire l'indirizzo del destinatario principale.\n"
     "4. Aggiungere un destinatario in CC.\n"
     "5. Inserire oggetto e corpo del messaggio.\n"
     "6. Cliccare 'Invia'.\n"
     "7. Scattare screenshot.\n"
     "8. Verificare la conferma di invio (assert).",
     "Il messaggio è inviato con successo con entrambi i destinatari (To e CC). Appare il toast di conferma."),

    (22, "Messaggi", "test_messaggi_20.py", "test_scarica_allegato_ricevuto",
     "Verifica il download di un allegato da un messaggio PEC ricevuto.",
     "1. Accedere con credenziali valide.\n"
     "2. Inviare a se stessi un messaggio con allegato PDF.\n"
     "3. Attendere la ricezione del messaggio.\n"
     "4. Aprire il messaggio ricevuto.\n"
     "5. Cliccare sul pulsante di download dell'allegato.\n"
     "6. Gestire il download tramite Playwright expect_download.\n"
     "7. Salvare il file nella cartella test-results.\n"
     "8. Scattare screenshot.\n"
     "9. Verificare che il file scaricato esista e abbia dimensione > 0 (assert).",
     "Il file allegato viene scaricato correttamente. Il file esiste nella directory di destinazione e ha dimensione maggiore di 0 byte."),

    (23, "Messaggi", "test_messaggi_21.py", "test_salva_e_usa_modello",
     "Verifica il salvataggio di un messaggio come modello e il suo successivo utilizzo.",
     "1. Accedere con credenziali valide.\n"
     "2. Cliccare 'Nuovo messaggio'.\n"
     "3. Inserire oggetto e corpo del messaggio modello.\n"
     "4. Salvare come modello invece di inviare.\n"
     "5. Navigare nella cartella Modelli.\n"
     "6. Verificare che il modello sia presente.\n"
     "7. Aprire il modello e inviarlo come nuovo messaggio.\n"
     "8. Scattare screenshot.",
     "Il modello è salvato e visibile nella cartella Modelli. Può essere aperto e utilizzato per inviare un nuovo messaggio."),

    (24, "Messaggi", "test_messaggi_22.py", "test_verifica_cartella_inviati",
     "Verifica che i messaggi inviati siano presenti nella cartella Inviati.",
     "1. Accedere con credenziali valide.\n"
     "2. Inviare un messaggio con oggetto univoco.\n"
     "3. Attendere la conferma di invio.\n"
     "4. Navigare nella cartella Inviati.\n"
     "5. Cercare il messaggio con l'oggetto univoco.\n"
     "6. Scattare screenshot.\n"
     "7. Verificare che il messaggio sia presente con il corretto oggetto (assert).",
     "Il messaggio inviato è visibile nella cartella Inviati con l'oggetto corretto."),

    (25, "Messaggi", "test_messaggi_23.py", "test_ricevuta_accettazione",
     "Verifica la ricezione della Ricevuta di Accettazione (RA) dopo l'invio di una PEC.",
     "1. Accedere con credenziali valide.\n"
     "2. Inviare un messaggio PEC con oggetto univoco al destinatario secondario.\n"
     "3. Navigare in In arrivo.\n"
     "4. Attendere l'arrivo della Ricevuta di Accettazione (15+ secondi).\n"
     "5. Aggiornare la vista messaggi.\n"
     "6. Cercare il messaggio con prefisso 'ACCETTAZIONE'.\n"
     "7. Scattare screenshot.\n"
     "8. Verificare la presenza della RA (assert).",
     "La Ricevuta di Accettazione (RA) è presente in In arrivo entro i tempi attesi. L'oggetto contiene il prefisso 'ACCETTAZIONE'."),

    (26, "Messaggi", "test_messaggi_24.py", "test_ricevuta_consegna",
     "Verifica la ricezione della Ricevuta di Consegna (RD) dopo l'invio di una PEC.",
     "1. Accedere con credenziali valide.\n"
     "2. Inviare un messaggio PEC con oggetto univoco al destinatario secondario.\n"
     "3. Navigare in In arrivo.\n"
     "4. Attendere l'arrivo della Ricevuta di Consegna (25+ secondi).\n"
     "5. Aggiornare la vista messaggi.\n"
     "6. Cercare il messaggio con prefisso 'CONSEGNA'.\n"
     "7. Scattare screenshot.\n"
     "8. Verificare la presenza della RD (assert).",
     "La Ricevuta di Consegna (RD) è presente in In arrivo entro i tempi attesi. L'oggetto contiene il prefisso 'CONSEGNA'."),

    (27, "Messaggi", "test_messaggi_25.py", "test_invio_allegati_multipli",
     "Verifica l'invio di un messaggio con allegati multipli.",
     "1. Accedere con credenziali valide.\n"
     "2. Cliccare 'Nuovo messaggio'.\n"
     "3. Inserire destinatario, oggetto e corpo.\n"
     "4. Cliccare il pulsante allegati e selezionare il primo file.\n"
     "5. Aggiungere un secondo allegato (stesso file o diverso).\n"
     "6. Verificare che almeno 2 allegati siano visibili nella finestra di composizione.\n"
     "7. Inviare il messaggio.\n"
     "8. Scattare screenshot.",
     "Almeno 2 allegati sono visibili nella finestra di composizione prima dell'invio. Il messaggio viene inviato con successo."),

    (28, "Messaggi", "test_messaggi_26.py", "test_selezione_multipla_batch",
     "Verifica la selezione multipla di messaggi e la comparsa della toolbar azioni batch.",
     "1. Accedere con credenziali valide.\n"
     "2. Inviare 2 messaggi a se stessi con oggetti univoci.\n"
     "3. Attendere la ricezione di entrambi in In arrivo.\n"
     "4. Passare il mouse sulla prima riga messaggio per rivelare il checkbox.\n"
     "5. Selezionare il checkbox della prima riga.\n"
     "6. Passare il mouse sulla seconda riga e selezionare il suo checkbox.\n"
     "7. Scattare screenshot.\n"
     "8. Verificare la comparsa della toolbar azioni batch (Elimina, Segna, ecc., assert).",
     "Dopo la selezione di almeno 2 messaggi, la toolbar delle azioni batch è visibile. I pulsanti batch (Elimina, Segna come letto, ecc.) sono presenti."),

    (29, "Messaggi", "test_messaggi_27.py", "test_anomalia_invio_non_pec",
     "Verifica che l'invio a un indirizzo non-PEC generi una notifica di anomalia.",
     "1. Accedere con credenziali valide.\n"
     "2. Cliccare 'Nuovo messaggio'.\n"
     "3. Inserire un indirizzo email non-PEC come destinatario (es. test@gmail.com).\n"
     "4. Inserire oggetto e corpo del messaggio.\n"
     "5. Inviare il messaggio.\n"
     "6. Attendere la ricezione della notifica di anomalia.\n"
     "7. Scattare screenshot.\n"
     "8. Verificare la presenza della notifica nell'inbox o come avviso UI (assert).",
     "Viene generata e ricevuta una notifica di anomalia per l'invio a un indirizzo non-PEC. L'oggetto contiene la parola 'anomalia'."),

    # ══════════════════════════════════════════════════════════════════════════
    # CALENDARIO
    # ══════════════════════════════════════════════════════════════════════════
    (30, "Calendario", "test_calendario_01.py", "test_creazione_evento_ricorrente",
     "Verifica la creazione di un evento ricorrente con regola di ricorrenza personalizzata.",
     "1. Accedere con credenziali valide.\n"
     "2. Navigare nella sezione Calendario.\n"
     "3. Cliccare 'Nuovo evento'.\n"
     "4. Inserire titolo e dettagli dell'evento.\n"
     "5. Abilitare la ricorrenza e configurare la regola (es. settimanale).\n"
     "6. Salvare l'evento.\n"
     "7. Verificare la comparsa e il testo del toast di conferma salvataggio (assert).\n"
     "8. Verificare la presenza dell'evento ricorrente nel calendario.\n"
     "9. Scattare screenshot.\n"
     "10. Eliminare l'evento (singolo o tutti, cleanup).",
     "L'evento ricorrente è creato e visibile nel calendario. Al momento dell'eliminazione è possibile scegliere di eliminare il singolo evento o tutti gli eventi della serie."),

    (31, "Calendario", "test_calendario_02.py", "test_creazione_modifica_evento",
     "Verifica la creazione e la modifica del titolo di un evento del calendario.",
     "1. Accedere con credenziali valide.\n"
     "2. Navigare nella sezione Calendario.\n"
     "3. Cliccare 'Nuovo evento'.\n"
     "4. Inserire titolo e dettagli.\n"
     "5. Salvare l'evento.\n"
     "6. Aprire l'evento appena creato.\n"
     "7. Modificare il titolo con uno nuovo.\n"
     "8. Salvare la modifica.\n"
     "9. Verificare la comparsa e il testo del toast di conferma salvataggio (assert).\n"
     "10. Verificare che il nuovo titolo sia visibile nel calendario.\n"
     "11. Scattare screenshot.\n"
     "12. Eliminare l'evento (cleanup).",
     "L'evento è creato con il titolo originale. Dopo la modifica, il calendario mostra il nuovo titolo. L'eliminazione completa il cleanup."),

    (32, "Calendario", "test_calendario_03.py", "test_creazione_invio_evento",
     "Verifica la creazione di un evento con invito al partecipante principale e secondario, il ricevimento dell'invito lato organizzatore e la ricezione effettiva lato destinatario secondario (secondo account, seconda pagina browser).",
     "1. Accedere con credenziali valide (account organizzatore).\n"
     "2. Navigare nella sezione Calendario.\n"
     "3. Creare un nuovo evento.\n"
     "4. Aggiungere il partecipante principale come invitato.\n"
     "5. Aggiungere anche il partecipante secondario come invitato (se distinto dal principale).\n"
     "6. Verificare il toast di conferma salvataggio evento (assert su testo IT/FR).\n"
     "7. Salvare e inviare l'invito.\n"
     "8. Navigare in In arrivo messaggi (account organizzatore) e verificare la ricevuta (assert).\n"
     "9. Aprire una seconda pagina/contesto browser, accedere con le credenziali del destinatario secondario.\n"
     "10. Navigare in In arrivo del destinatario secondario e attendere (polling) l'arrivo dell'email di invito.\n"
     "11. Verificare che l'invito sia effettivamente arrivato nella inbox del destinatario secondario (assert).\n"
     "12. Eliminare l'evento (cleanup lato organizzatore).",
     "L'evento è creato con entrambi i partecipanti invitati. L'invito viene inviato come messaggio PEC ed è visibile sia nella ricevuta dell'organizzatore sia, effettivamente, nella inbox del destinatario secondario tramite un secondo account/pagina browser."),

    (33, "Calendario", "test_calendario_04.py", "test_import_export_calendario",
     "Verifica l'importazione di un calendario da file ICS e il tentativo di esportazione.",
     "1. Accedere con credenziali valide.\n"
     "2. Navigare nella sezione Calendario.\n"
     "3. Creare un nuovo evento di test.\n"
     "4. Cliccare il pulsante 'Importa'.\n"
     "5. Selezionare il file ICS tramite il file picker.\n"
     "6. Confermare l'importazione.\n"
     "7. Tentare l'esportazione del calendario (via pulsante o menu contestuale).\n"
     "8. Scattare screenshot.\n"
     "9. Eliminare gli eventi importati (cleanup).",
     "Il file ICS è importato senza errori. Il calendario mostra gli eventi importati. L'esportazione (se disponibile) genera un file ICS."),

    (34, "Calendario", "test_calendario_05.py", "test_creazione_calendario",
     "Verifica la creazione di un nuovo calendario personalizzato con nome e colore.",
     "1. Accedere con credenziali valide.\n"
     "2. Navigare nella sezione Calendario.\n"
     "3. Aprire le opzioni per aggiungere un nuovo calendario.\n"
     "4. Inserire nome e selezionare colore del calendario.\n"
     "5. Confermare la creazione.\n"
     "6. Verificare la comparsa del toast di conferma.\n"
     "7. Verificare che il nuovo calendario sia visibile nella sidebar.\n"
     "8. Scattare screenshot.\n"
     "9. Eliminare il calendario creato (cleanup).",
     "Il nuovo calendario è creato con successo e appare nella sidebar. Il toast di conferma è visibile. Il calendario può essere eliminato."),

    (35, "Calendario", "test_calendario_06.py", "test_cambio_vista_calendario",
     "Verifica il cambio tra le diverse viste del calendario (Giorno, Settimana, Mese, Eventi).",
     "1. Accedere con credenziali valide.\n"
     "2. Navigare nella sezione Calendario.\n"
     "3. Cliccare il pulsante vista 'Giorno'.\n"
     "4. Scattare screenshot della vista Giorno.\n"
     "5. Cliccare il pulsante vista 'Mese'.\n"
     "6. Scattare screenshot della vista Mese.\n"
     "7. Cliccare il pulsante vista 'Eventi'.\n"
     "8. Scattare screenshot della vista Eventi.\n"
     "9. Scattare screenshot finale.",
     "Il calendario cambia correttamente tra le viste Giorno, Mese ed Eventi. Ogni vista mostra il layout corretto e screenshot vengono salvati."),

    (36, "Calendario", "test_calendario_07.py", "test_naviga_periodi_calendario",
     "Verifica la navigazione tra periodi del calendario tramite i pulsanti Precedente e Successivo.",
     "1. Accedere con credenziali valide.\n"
     "2. Navigare nella sezione Calendario.\n"
     "3. Leggere il contenuto corrente dell'header del calendario.\n"
     "4. Cliccare il pulsante 'Successivo' (>).\n"
     "5. Verificare che l'header sia cambiato rispetto al valore iniziale.\n"
     "6. Cliccare il pulsante 'Precedente' (<).\n"
     "7. Scattare screenshot.\n"
     "8. Verificare il ritorno al periodo iniziale (assert).",
     "L'header del calendario cambia correttamente dopo la navigazione avanti/indietro tra i periodi."),

    (37, "Calendario", "test_calendario_08.py", "test_evento_tutto_il_giorno",
     "Verifica la creazione di un evento contrassegnato come 'Giornata intera'.",
     "1. Accedere con credenziali valide.\n"
     "2. Navigare nella sezione Calendario.\n"
     "3. Cliccare 'Nuovo evento'.\n"
     "4. Inserire il titolo dell'evento.\n"
     "5. Abilitare il toggle/checkbox 'Giornata intera'.\n"
     "6. Salvare l'evento.\n"
     "7. Verificare la presenza dell'evento nel calendario come evento giornaliero.\n"
     "8. Scattare screenshot.\n"
     "9. Eliminare l'evento (cleanup).",
     "L'evento è creato come evento 'Giornata intera' ed è visibile nella banda dedicata agli eventi a tutto giorno nel calendario."),

    (38, "Calendario", "test_calendario_09.py", "test_evento_con_promemoria",
     "Verifica la configurazione e il salvataggio di un promemoria per un evento del calendario.",
     "1. Accedere con credenziali valide.\n"
     "2. Navigare nella sezione Calendario.\n"
     "3. Cliccare 'Nuovo evento'.\n"
     "4. Inserire il titolo dell'evento.\n"
     "5. Configurare un promemoria (es. 15 minuti prima).\n"
     "6. Salvare l'evento.\n"
     "7. Riaprire l'evento.\n"
     "8. Verificare che il promemoria sia presente e configurato.\n"
     "9. Scattare screenshot.\n"
     "10. Eliminare l'evento (cleanup).",
     "Il promemoria è configurato e salvato correttamente nell'evento. Alla riapertura dell'evento il promemoria è visibile con la corretta configurazione."),

    (39, "Calendario", "test_calendario_10.py", "test_ricerca_nel_calendario",
     "Verifica la funzione di ricerca eventi nel calendario.",
     "1. Accedere con credenziali valide.\n"
     "2. Navigare nella sezione Calendario.\n"
     "3. Creare un evento con titolo univoco.\n"
     "4. Salvare l'evento.\n"
     "5. Verificare la comparsa e il testo del toast di conferma salvataggio (assert).\n"
     "6. Aprire la barra di ricerca nel calendario.\n"
     "7. Inserire il titolo univoco dell'evento.\n"
     "8. Avviare la ricerca.\n"
     "9. Verificare che l'evento sia trovato nei risultati.\n"
     "10. Scattare screenshot.\n"
     "11. Eliminare l'evento (cleanup).",
     "L'evento con il titolo univoco è trovato e visualizzato nei risultati di ricerca del calendario."),

    # ══════════════════════════════════════════════════════════════════════════
    # CONTATTI
    # ══════════════════════════════════════════════════════════════════════════
    (40, "Contatti", "test_contatti_01.py", "test_aggiungere_nuovo_contatto",
     "Verifica la creazione di un nuovo contatto nella rubrica della casella PEC.",
     "1. Accedere con credenziali valide.\n"
     "2. Navigare nella sezione Contatti.\n"
     "3. Cliccare 'Nuovo contatto'.\n"
     "4. Inserire nome, cognome e indirizzo email.\n"
     "5. Salvare il contatto.\n"
     "6. Verificare la presenza del contatto nella rubrica.\n"
     "7. Scattare screenshot.\n"
     "8. Eliminare il contatto (cleanup).",
     "Il contatto è creato e visibile nella rubrica con nome, cognome e email corretti. Può essere eliminato correttamente."),

    (41, "Contatti", "test_contatti_02.py", "test_aggiungere_nuovo_gruppo",
     "Verifica la creazione di un gruppo contatti e l'aggiunta di contatti al gruppo.",
     "1. Accedere con credenziali valide.\n"
     "2. Navigare nella sezione Contatti.\n"
     "3. Cliccare 'Nuovo gruppo'.\n"
     "4. Inserire il nome del gruppo.\n"
     "5. Salvare il gruppo.\n"
     "6. Aggiungere uno o più contatti al gruppo.\n"
     "7. Verificare che il gruppo sia visibile nella sidebar.\n"
     "8. Scattare screenshot.\n"
     "9. Verificare che i contatti siano nel gruppo (assert).",
     "Il gruppo è creato e visibile nella sidebar. I contatti aggiunti sono visibili nel gruppo."),

    (42, "Contatti", "test_contatti_03.py", "test_esporta_contatti",
     "Verifica l'esportazione della rubrica contatti in formato CSV.",
     "1. Accedere con credenziali valide.\n"
     "2. Navigare nella sezione Contatti.\n"
     "3. Aprire il menu Impostazioni/Opzioni della rubrica.\n"
     "4. Selezionare 'Esporta' o 'Esporta come CSV'.\n"
     "5. Gestire il download del file tramite Playwright expect_download.\n"
     "6. Verificare che il file CSV sia scaricato.\n"
     "7. Scattare screenshot.\n"
     "8. Verificare che il file abbia dimensione > 0 (assert).",
     "Il file CSV della rubrica è scaricato correttamente. Il file esiste nella directory di destinazione e ha dimensione maggiore di 0 byte."),

    (43, "Contatti", "test_contatti_04.py", "test_importa_contatti",
     "Verifica l'importazione di contatti da un file CSV nella rubrica PEC.",
     "1. Accedere con credenziali valide.\n"
     "2. Navigare nella sezione Contatti.\n"
     "3. Aprire il menu Impostazioni/Opzioni della rubrica.\n"
     "4. Selezionare 'Importa'.\n"
     "5. Selezionare il file CSV di test tramite il file picker.\n"
     "6. Confermare l'importazione.\n"
     "7. Verificare il messaggio di successo.\n"
     "8. Scattare screenshot.\n"
     "9. Verificare che i contatti importati siano nella rubrica (assert).",
     "I contatti del file CSV sono importati con successo nella rubrica. Il messaggio di conferma è visibile."),

    (44, "Contatti", "test_contatti_05.py", "test_modifica_contatto",
     "Verifica la modifica dei dati di un contatto esistente nella rubrica.",
     "1. Accedere con credenziali valide.\n"
     "2. Navigare nella sezione Contatti.\n"
     "3. Cercare il contatto da modificare.\n"
     "4. Aprire il contatto.\n"
     "5. Cliccare 'Modifica'.\n"
     "6. Cambiare il cognome con un valore nuovo univoco.\n"
     "7. Salvare le modifiche.\n"
     "8. Cercare il contatto con il nuovo cognome.\n"
     "9. Scattare screenshot.\n"
     "10. Verificare che il contatto sia trovato con i dati aggiornati (assert).",
     "Il contatto è modificato correttamente. La ricerca con il nuovo cognome restituisce il contatto aggiornato."),

    (45, "Contatti", "test_contatti_06.py", "test_elimina_contatto",
     "Verifica l'eliminazione di un contatto dalla rubrica.",
     "1. Accedere con credenziali valide.\n"
     "2. Navigare nella sezione Contatti.\n"
     "3. Creare un contatto di test con nome univoco.\n"
     "4. Selezionare il contatto.\n"
     "5. Cliccare 'Elimina contatto'.\n"
     "6. Confermare l'eliminazione.\n"
     "7. Cercare il contatto eliminato.\n"
     "8. Scattare screenshot.\n"
     "9. Verificare che non sia più presente nella rubrica (assert).",
     "Il contatto eliminato non è più presente nella rubrica. La ricerca per nome non restituisce risultati."),

    (46, "Contatti", "test_contatti_07.py", "test_ricerca_contatto",
     "Verifica la ricerca di un contatto nella rubrica tramite la barra di ricerca.",
     "1. Accedere con credenziali valide.\n"
     "2. Navigare nella sezione Contatti.\n"
     "3. Creare un contatto con nome univoco.\n"
     "4. Inserire il nome univoco nella barra di ricerca.\n"
     "5. Avviare la ricerca.\n"
     "6. Verificare che il contatto cercato sia nei risultati.\n"
     "7. Scattare screenshot.\n"
     "8. Eliminare il contatto (cleanup).",
     "Il contatto con il nome univoco è trovato e visibile nei risultati di ricerca della rubrica."),

    (47, "Contatti", "test_contatti_08.py", "test_contatto_preferito",
     "Verifica l'aggiunta di un contatto ai preferiti e la sua visualizzazione nella sezione Preferiti.",
     "1. Accedere con credenziali valide.\n"
     "2. Navigare nella sezione Contatti.\n"
     "3. Selezionare un contatto dalla rubrica.\n"
     "4. Cliccare il pulsante stella (preferito).\n"
     "5. Navigare nella sezione 'Contatti preferiti' nella sidebar.\n"
     "6. Verificare che il contatto sia presente nella sezione Preferiti.\n"
     "7. Scattare screenshot.\n"
     "8. Rimuovere il contatto dai preferiti (cleanup).",
     "Il contatto aggiunto ai preferiti è visibile nella sezione Contatti preferiti nella sidebar."),

    (48, "Contatti", "test_contatti_09.py", "test_scrivi_email_da_contatto",
     "Verifica l'apertura della finestra di composizione email direttamente da un contatto.",
     "1. Accedere con credenziali valide.\n"
     "2. Navigare nella sezione Contatti.\n"
     "3. Selezionare un contatto dalla rubrica.\n"
     "4. Cliccare il pulsante 'Scrivi email' o 'Invia email'.\n"
     "5. Verificare che la finestra di composizione si apra.\n"
     "6. Scattare screenshot.\n"
     "7. Verificare che il campo destinatario sia pre-compilato con l'email del contatto (assert).",
     "La finestra di composizione si apre correttamente con il campo destinatario pre-compilato con l'indirizzo email del contatto selezionato."),

    (49, "Contatti", "test_contatti_10.py", "test_elimina_gruppo",
     "Verifica la creazione e l'eliminazione di un gruppo contatti.",
     "1. Accedere con credenziali valide.\n"
     "2. Navigare nella sezione Contatti.\n"
     "3. Creare un nuovo gruppo con nome univoco.\n"
     "4. Verificare la presenza del gruppo nella sidebar.\n"
     "5. Selezionare il gruppo.\n"
     "6. Cliccare 'Elimina gruppo'.\n"
     "7. Confermare l'eliminazione.\n"
     "8. Scattare screenshot.\n"
     "9. Verificare che il gruppo non sia più nella sidebar (assert).",
     "Il gruppo è eliminato correttamente e non è più visibile nella sidebar dei contatti."),

    # ══════════════════════════════════════════════════════════════════════════
    # IMPOSTAZIONI
    # ══════════════════════════════════════════════════════════════════════════
    (50, "Impostazioni", "test_impostazioni_01.py", "test_crea_firma",
     "Verifica la creazione di una firma email nella sezione Impostazioni.",
     "1. Accedere con credenziali valide.\n"
     "2. Navigare nella sezione Impostazioni.\n"
     "3. Aprire la sotto-sezione 'Firme'.\n"
     "4. Cliccare 'Nuova firma'.\n"
     "5. Inserire il testo della firma.\n"
     "6. Salvare la firma.\n"
     "7. Verificare che la firma sia visibile nell'elenco firme.\n"
     "8. Scattare screenshot.\n"
     "9. Eliminare la firma (cleanup).",
     "La firma è creata e visibile nell'elenco firme. Può essere eliminata correttamente."),

    (51, "Impostazioni", "test_impostazioni_02.py", "test_riquadro_di_lettura",
     "Verifica la modifica della posizione del riquadro di lettura nei messaggi.",
     "1. Accedere con credenziali valide.\n"
     "2. Navigare nella sezione Impostazioni > Posta.\n"
     "3. Individuare l'opzione 'Riquadro di lettura'.\n"
     "4. Selezionare l'opzione 'A destra'.\n"
     "5. Salvare le impostazioni.\n"
     "6. Verificare l'applicazione della modifica.\n"
     "7. Selezionare l'opzione 'In basso'.\n"
     "8. Scattare screenshot.\n"
     "9. Verificare l'applicazione (assert).",
     "Le opzioni del riquadro di lettura sono disponibili e selezionabili. La modifica viene salvata e applicata correttamente."),

    (52, "Impostazioni", "test_impostazioni_03.py", "test_regole_messaggi",
     "Verifica l'accesso e la creazione di regole per i messaggi nelle Impostazioni.",
     "1. Accedere con credenziali valide.\n"
     "2. Navigare nella sezione Impostazioni.\n"
     "3. Aprire la sotto-sezione 'Regole messaggi' o 'Filtri'.\n"
     "4. Cliccare 'Nuova regola'.\n"
     "5. Configurare le condizioni della regola (mittente, oggetto, ecc.).\n"
     "6. Configurare l'azione (sposta in cartella, elimina, ecc.).\n"
     "7. Salvare la regola.\n"
     "8. Scattare screenshot.\n"
     "9. Verificare la presenza della regola nell'elenco (assert).",
     "La sezione Regole messaggi è accessibile. Il form di creazione nuova regola si apre correttamente."),

    (53, "Impostazioni", "test_impostazioni_04.py", "test_risposta_automatica",
     "Verifica l'accesso alla sezione Avvisi e report nelle Impostazioni.",
     "1. Accedere con credenziali valide.\n"
     "2. Navigare nella sezione Impostazioni.\n"
     "3. Aprire la sotto-sezione 'Avvisi e report' o 'Risposta automatica'.\n"
     "4. Verificare che la pagina sia caricata correttamente.\n"
     "5. Scattare screenshot.\n"
     "6. Verificare la presenza dei controlli di configurazione (assert).",
     "La sezione Avvisi e report è accessibile e la pagina si carica correttamente con i controlli di configurazione visibili."),

    (54, "Impostazioni", "test_impostazioni_05.py", "test_posta_indesiderata",
     "Verifica la sezione Posta indesiderata e il blocco di un mittente.",
     "1. Accedere con credenziali valide.\n"
     "2. Navigare nella sezione Impostazioni.\n"
     "3. Aprire la sotto-sezione 'Posta indesiderata'.\n"
     "4. Aggiungere un indirizzo email alla lista di blocco.\n"
     "5. Salvare le impostazioni.\n"
     "6. Verificare che l'indirizzo sia nella lista di blocco.\n"
     "7. Scattare screenshot.\n"
     "8. Rimuovere l'indirizzo dalla lista (cleanup).",
     "La sezione Posta indesiderata è accessibile. È possibile aggiungere e rimuovere indirizzi dalla lista di blocco."),

    (55, "Impostazioni", "test_impostazioni_06.py", "test_impostazioni_cestino",
     "Verifica le opzioni di configurazione del Cestino nelle Impostazioni.",
     "1. Accedere con credenziali valide.\n"
     "2. Navigare nella sezione Impostazioni.\n"
     "3. Aprire la sotto-sezione 'Cestino'.\n"
     "4. Verificare la presenza delle opzioni di configurazione (svuotamento automatico, ecc.).\n"
     "5. Modificare un'impostazione.\n"
     "6. Salvare.\n"
     "7. Scattare screenshot.\n"
     "8. Verificare il salvataggio (assert).",
     "La sezione Cestino nelle Impostazioni è accessibile e le opzioni di configurazione sono presenti e modificabili."),

    (56, "Impostazioni", "test_impostazioni_07.py", "test_storico_accessi",
     "Verifica l'accesso e la visualizzazione dello storico degli accessi alla casella PEC.",
     "1. Accedere con credenziali valide.\n"
     "2. Navigare nella sezione Impostazioni.\n"
     "3. Aprire la sotto-sezione 'Storico accessi' o 'Accessi recenti'.\n"
     "4. Verificare che la pagina si carichi correttamente.\n"
     "5. Verificare la presenza del pulsante/link per visualizzare lo storico.\n"
     "6. Scattare screenshot.\n"
     "7. Cliccare il pulsante e verificare la visualizzazione della lista accessi (assert).",
     "La sezione Storico accessi è accessibile. Il pulsante per visualizzare lo storico è visibile e funzionante."),

    # ══════════════════════════════════════════════════════════════════════════
    # ARCHIVIO
    # ══════════════════════════════════════════════════════════════════════════
    (57, "Archivio", "test_archivio_01.py", "test_configurazione_archivio",
     "Verifica la pagina di configurazione archivio: struttura, radio button e checkbox opzioni. "
     "Modifica la selezione, salva e verifica il persist. Skip se la feature non e' disponibile nell'ambiente.",
     "1. Accedere con credenziali valide.\n"
     "2. Navigare in Impostazioni > Home e verificare la presenza del pulsante 'Archivio'.\n"
     "3. Skip se la feature non e' disponibile (pytest.skip).\n"
     "4. Navigare direttamente a /new/settings/archive.\n"
     "5. Verificare h1 'Archivio' e h2 'Configurazione Archivio' visibili.\n"
     "6. Verificare la presenza di almeno 2 radio button.\n"
     "7. Verificare la presenza delle opzioni checkbox (RA, RD, PEC ricevute/inviate).\n"
     "8. Selezionare il secondo radio button ('Scegli quali messaggi archiviare').\n"
     "9. Abilitare la prima checkbox visibile.\n"
     "10. Salvare con il pulsante primario.\n"
     "11. Verificare il toast di conferma salvataggio.\n"
     "12. Ricaricare la pagina e verificare che la selezione sia mantenuta.\n"
     "13. Scattare screenshot.\n"
     "14. Cleanup: ripristinare il primo radio button e salvare.",
     "La pagina di configurazione Archivio mostra la struttura attesa. Il salvataggio di una configurazione personalizzata persiste dopo reload. Il cleanup ripristina lo stato iniziale."),

    (58, "Archivio", "test_archivio_02.py", "test_archivio_messaggio_inviato",
     "Verifica che i messaggi inviati/ricevuti vengano archiviati secondo la configurazione impostata. "
     "Imposta 'Archivia tutti', invia un messaggio a se stesso, poi verifica la presenza del messaggio in Archivio.",
     "1. Accedere con credenziali valide.\n"
     "2. Verificare la disponibilita' della feature Archivio (pytest.skip se assente).\n"
     "3. Navigare a /new/settings/archive e selezionare il primo radio button ('Archivia tutti').\n"
     "4. Salvare la configurazione.\n"
     "5. Creare e inviare un messaggio PEC a se stessi con oggetto univoco (timestamp).\n"
     "6. Attendere 10 secondi per permettere l'archiviazione.\n"
     "7. Navigare a /new/messages/INBOX.\n"
     "8. Aprire la sezione Archivio tramite il menu waffle (9 punti nell'header).\n"
     "9. Verificare che il messaggio con l'oggetto univoco sia presente in Archivio.\n"
     "10. Scattare screenshot.\n"
     "11. Fallback: se il waffle non funziona, verificare l'accesso via settings URL (assert).",
     "Il messaggio inviato a se stessi compare nella sezione Archivio. La configurazione 'Archivia tutti' funziona correttamente."),

    # ══════════════════════════════════════════════════════════════════════════
    # CONSERVAZIONE
    # ══════════════════════════════════════════════════════════════════════════
    (59, "Conservazione", "test_conservazione_01.py", "test_cartella_da_conservare",
     "Verifica l'accesso alla cartella 'Da conservare' nella sidebar e la struttura della pagina. "
     "Skip se la feature non e' disponibile nell'ambiente.",
     "1. Accedere con credenziali valide.\n"
     "2. Navigare a /new/messages/INBOX.\n"
     "3. Verificare la presenza del pulsante 'Da conservare' nella sidebar.\n"
     "4. Skip se non visibile (pytest.skip).\n"
     "5. Cliccare il pulsante 'Da conservare'.\n"
     "6. Verificare che l'URL contenga 'CONSERVAZIONE'.\n"
     "7. Verificare la presenza di testo informativo sulla conservazione.\n"
     "8. Scattare screenshot.\n"
     "9. Verificare la presenza del link alle Impostazioni (assert).",
     "La cartella 'Da conservare' e' accessibile dalla sidebar. L'URL contiene 'CONSERVAZIONE' e la pagina mostra contenuti informativi sulla conservazione a norma."),

    (60, "Conservazione", "test_conservazione_02.py", "test_conservazione_da_waffle",
     "Verifica l'accesso alla sezione Conservazione tramite il menu waffle (9 punti nell'header). "
     "Fallback su navigazione diretta se il waffle non e' disponibile.",
     "1. Accedere con credenziali valide.\n"
     "2. Navigare a /new/messages/INBOX.\n"
     "3. Verificare la disponibilita' della feature Conservazione (pytest.skip se assente).\n"
     "4. Cercare e cliccare il menu waffle (multi-fallback selectors).\n"
     "5. Cercare la voce 'Conservazione' nel dropdown del waffle.\n"
     "6. Cliccare la voce (gestire eventuale apertura in nuova tab).\n"
     "7. Verificare che la pagina di Conservazione si carichi correttamente.\n"
     "8. Scattare screenshot.\n"
     "9. Fallback: se il waffle non funziona, navigare direttamente alla URL di Conservazione e verificare (assert).",
     "La sezione Conservazione e' accessibile tramite il menu waffle. La pagina di Conservazione si carica e mostra i contenuti attesi."),

    # ══════════════════════════════════════════════════════════════════════════
    # FATTURE
    # ══════════════════════════════════════════════════════════════════════════
    (61, "Fatture", "test_fatture_01.py", "test_visualizza_fattura_importata",
     "Verifica l'importazione e la visualizzazione di una fattura elettronica FatturaPA (EML da SDI) "
     "nella sezione 'Fatture ricevute'. Skip se la feature non e' disponibile.",
     "1. Accedere con credenziali valide.\n"
     "2. Verificare la presenza del pulsante 'Fatture ricevute' nella sidebar.\n"
     "3. Skip se non visibile (pytest.skip).\n"
     "4. Cliccare 'Fatture ricevute' e verificare che l'URL contenga 'ArubaVrtSearch/Fatturazione%20Elettronica'.\n"
     "5. Tornare a INBOX.\n"
     "6. Aprire 'Gestione messaggi' nella sidebar.\n"
     "7. Cliccare 'Importa'.\n"
     "8. Attendere il file input (#hidden_input) e impostare il file EML di test.\n"
     "9. Attendere il messaggio di conferma importazione.\n"
     "10. Navigare a 'Fatture ricevute'.\n"
     "11. Scattare screenshot.\n"
     "12. Verificare che sia presente almeno un messaggio con riferimento a SDI o FatturaPA (assert).",
     "La fattura EML viene importata correttamente. La sezione 'Fatture ricevute' mostra il messaggio importato proveniente dall'SDI."),

    (62, "Fatture", "test_fatture_02.py", "test_impostazioni_leggi_fatture",
     "Verifica la sezione 'Leggi fatture' nelle Impostazioni: struttura pagina e contenuto informativo. "
     "Skip se la feature non e' disponibile nell'ambiente.",
     "1. Accedere con credenziali valide.\n"
     "2. Navigare a /new/settings/home.\n"
     "3. Verificare la presenza del pulsante 'Leggi fatture'.\n"
     "4. Skip se non visibile (pytest.skip).\n"
     "5. Navigare direttamente a /new/settings/read-invoices.\n"
     "6. Verificare che la pagina si carichi (stato domcontentloaded).\n"
     "7. Verificare la presenza dell'h1 o testo 'Leggi fatture'.\n"
     "8. Scattare screenshot.\n"
     "9. Verificare la presenza di contenuto informativo (cassetto fiscale, AdE, ecc., assert).",
     "La sezione 'Leggi fatture' e' accessibile nelle Impostazioni. La pagina mostra informazioni sul servizio di collegamento al cassetto fiscale dell'Agenzia delle Entrate."),

    # ══════════════════════════════════════════════════════════════════════════
    # ACCESSI ALTRI ACCOUNT
    # ══════════════════════════════════════════════════════════════════════════
    (63, "Accessi", "test_accessi_01.py", "test_panoramica_accessi",
     "Verifica la pagina di panoramica 'Accessi altri account': struttura, card Multiutente PEC "
     "e Supervisore360, e navigazione alle sotto-sezioni. Skip se la feature non e' disponibile.",
     "1. Accedere con credenziali valide.\n"
     "2. Navigare a /new/settings/home e verificare la presenza di 'Accessi altri account'.\n"
     "3. Skip se non visibile (pytest.skip).\n"
     "4. Navigare a /new/settings/other-accounts/overview.\n"
     "5. Verificare h1 'Accessi altri account'.\n"
     "6. Verificare la presenza delle card 'Multiutente PEC' e 'Supervisore360'.\n"
     "7. Verificare la presenza di almeno 2 pulsanti 'Gestisci'.\n"
     "8. Navigare alla sotto-sezione Multiutente PEC e verificare il contenuto.\n"
     "9. Tornare alla panoramica e navigare a Supervisore360.\n"
     "10. Scattare screenshot.\n"
     "11. Verificare il contenuto della sezione Supervisore360 (assert).",
     "La pagina di panoramica 'Accessi altri account' mostra le card Multiutente PEC e Supervisore360 con i relativi pulsanti 'Gestisci'. La navigazione alle sotto-sezioni funziona correttamente."),

    (64, "Accessi", "test_accessi_02.py", "test_form_multiutente_pec",
     "Verifica la sezione Multiutente PEC: apertura del form di aggiunta utente e presenza "
     "dei campi. Non invia il form per evitare side-effect. Skip se la feature non e' disponibile.",
     "1. Accedere con credenziali valide.\n"
     "2. Verificare la disponibilita' della feature 'Accessi altri account' (pytest.skip se assente).\n"
     "3. Navigare a /new/settings/other-accounts/overview.\n"
     "4. Cliccare il pulsante 'Multiutente PEC' nella sidebar o il pulsante 'Gestisci'.\n"
     "5. Verificare il contenuto della pagina Multiutente PEC.\n"
     "6. Cercare e cliccare il pulsante 'Aggiungi' (multi-fallback selectors).\n"
     "7. Skip se il pulsante non e' trovato (pytest.skip).\n"
     "8. Verificare che il form o dialog si sia aperto con campi input.\n"
     "9. Scattare screenshot.\n"
     "10. Chiudere il form senza inviare (tasto Annulla o Escape).",
     "Il form di aggiunta Multiutente PEC si apre correttamente e mostra i campi di input. La chiusura senza invio non produce side-effect."),

    (65, "Accessi", "test_accessi_03.py", "test_form_supervisore360",
     "Verifica la sezione Supervisore360: apertura del form di aggiunta delega e presenza "
     "del campo email/delegato. Non invia il form. Skip se la feature non e' disponibile.",
     "1. Accedere con credenziali valide.\n"
     "2. Verificare la disponibilita' della feature 'Accessi altri account' (pytest.skip se assente).\n"
     "3. Navigare a /new/settings/other-accounts/overview.\n"
     "4. Cliccare il pulsante 'Supervisore360' nella sidebar.\n"
     "5. Skip se non visibile (pytest.skip).\n"
     "6. Verificare il contenuto della pagina Supervisore360.\n"
     "7. Cercare e cliccare il pulsante 'Aggiungi' o 'Nuova delega' (multi-fallback selectors).\n"
     "8. Skip se il pulsante non e' trovato (pytest.skip).\n"
     "9. Verificare che il form o dialog si sia aperto con campi email/input.\n"
     "10. Scattare screenshot.\n"
     "11. Chiudere il form senza inviare (tasto Annulla o Escape).",
     "Il form di aggiunta delega Supervisore360 si apre e mostra i campi input attesi. La chiusura senza invio non produce side-effect reali (nessuna delega viene inviata)."),
]

# ── MODULE CONFIG (header color, row-even color) ──────────────────────────────
MODULE_STYLE = {
    "Login":          (C_MID_BLUE,   C_LIGHT_BLUE),
    "Messaggi":       (C_GREEN,      C_LIGHT_GREEN),
    "Calendario":     (C_PURPLE,     C_LIGHT_PUR),
    "Contatti":       (C_ORANGE,     C_LIGHT_ORA),
    "Impostazioni":   (C_TEAL,       C_LIGHT_TEA),
    "Archivio":       (C_MAROON,     C_LIGHT_MAR),
    "Conservazione":  (C_INDIGO,     C_LIGHT_IND),
    "Fatture":        (C_OLIVE,      C_LIGHT_OLI),
    "Accessi":        (C_PLUM,       C_LIGHT_PLU),
}

# ── WORKBOOK SETUP ────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

# ── SHEET: INDICE ─────────────────────────────────────────────────────────────
ws_idx = wb.active
ws_idx.title = "Indice"
ws_idx.sheet_view.showGridLines = False
ws_idx.column_dimensions["A"].width = 5
ws_idx.column_dimensions["B"].width = 22
ws_idx.column_dimensions["C"].width = 10
ws_idx.column_dimensions["D"].width = 55

# Title
ws_idx.merge_cells("A1:D1")
title_cell = ws_idx["A1"]
title_cell.value = "SUITE DI TEST – WEBMAIL PEC ARUBA"
title_cell.font = Font(name="Calibri", bold=True, size=18, color=C_WHITE)
title_cell.fill = fill(C_DARK_BLUE)
title_cell.alignment = align("center", "center", False)
ws_idx.row_dimensions[1].height = 42

ws_idx.merge_cells("A2:D2")
sub_cell = ws_idx["A2"]
sub_cell.value = "Documentazione completa dei casi di test automatizzati"
sub_cell.font = Font(name="Calibri", italic=True, size=11, color="808080")
sub_cell.fill = fill("EEF4FB")
sub_cell.alignment = align("center", "center", False)
ws_idx.row_dimensions[2].height = 22

ws_idx.row_dimensions[3].height = 10

# Headers
headers_idx = ["#", "Modulo", "Totale Test", "Descrizione Modulo"]
for col, h in enumerate(headers_idx, 1):
    c = ws_idx.cell(row=4, column=col)
    c.value = h
    c.font = Font(name="Calibri", bold=True, size=11, color=C_WHITE)
    c.fill = fill(C_DARK_BLUE)
    c.alignment = align("center", "center", False)
    c.border = BORDER_THIN
ws_idx.row_dimensions[4].height = 22

MODULE_DESCRIPTIONS = {
    "Login":         "Test di autenticazione: verifica login con credenziali errate e funzione di logout.",
    "Messaggi":      "Test sulla gestione dei messaggi PEC: invio, ricezione, risposta, allegati, cartelle, etichette, ricevute (RA/RD) e funzioni avanzate.",
    "Calendario":    "Test sul modulo Calendario: creazione/modifica eventi, ricorrenze, importazione ICS, cambio viste e navigazione periodi.",
    "Contatti":      "Test sulla rubrica contatti: aggiunta, modifica, eliminazione, ricerca, gruppi, preferiti, import/export CSV.",
    "Impostazioni":  "Test sulle impostazioni della casella: firme, riquadro di lettura, regole, posta indesiderata, cestino e storico accessi.",
    "Archivio":      "Test sulla funzionalita' Archivio: configurazione delle regole di archiviazione e verifica dell'archiviazione effettiva dei messaggi inviati/ricevuti.",
    "Conservazione": "Test sulla Conservazione a norma: accesso alla cartella 'Da conservare' nella sidebar e tramite menu waffle.",
    "Fatture":       "Test sulle fatture elettroniche: importazione EML da SDI e visualizzazione in 'Fatture ricevute'; configurazione sezione 'Leggi fatture' nelle Impostazioni.",
    "Accessi":       "Test su 'Accessi altri account': panoramica Multiutente PEC e Supervisore360, apertura form di aggiunta utente/delega senza submit.",
}

from collections import Counter
module_counts = Counter(t[1] for t in TESTS)

for i, (mod, desc) in enumerate(MODULE_DESCRIPTIONS.items()):
    row = 5 + i
    bg_h, bg_even = MODULE_STYLE[mod]
    row_data = [i + 1, mod, module_counts[mod], desc]
    bg = bg_even if i % 2 == 0 else C_LIGHT_GRAY
    for col, val in enumerate(row_data, 1):
        c = ws_idx.cell(row=row, column=col)
        c.value = val
        c.font = Font(name="Calibri", size=10, color="1A1A1A",
                      bold=(col == 2))
        c.fill = fill(bg)
        c.alignment = align("center" if col in (1, 3) else "left", "center", col == 4)
        c.border = BORDER_THIN
    ws_idx.row_dimensions[row].height = 28

# Total row  (row 4 = headers, row 5..5+n-1 = modules, row 5+n = total)
total_row = 4 + len(MODULE_DESCRIPTIONS) + 1
ws_idx.merge_cells(f"A{total_row}:B{total_row}")
tot1 = ws_idx[f"A{total_row}"]
tot1.value = "TOTALE CASI DI TEST"
tot1.font = Font(name="Calibri", bold=True, size=11, color=C_WHITE)
tot1.fill = fill(C_DARK_BLUE)
tot1.alignment = align("center", "center", False)
tot1.border = BORDER_THIN

tot2 = ws_idx[f"C{total_row}"]
tot2.value = len(TESTS)
tot2.font = Font(name="Calibri", bold=True, size=12, color=C_GOLD)
tot2.fill = fill(C_DARK_BLUE)
tot2.alignment = align("center", "center", False)
tot2.border = BORDER_THIN

ws_idx[f"D{total_row}"].fill = fill(C_DARK_BLUE)
ws_idx[f"D{total_row}"].border = BORDER_THIN


# ── SHEET: TUTTI I TEST ────────────────────────────────────────────────────────
ws = wb.create_sheet(title="Suite Completa")
ws.sheet_view.showGridLines = False

# Column widths
col_widths = [6, 14, 30, 40, 60, 50, 60]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Title row
ws.merge_cells("A1:G1")
t = ws["A1"]
t.value = "SUITE DI TEST – WEBMAIL PEC ARUBA  |  Casi di Test Automatizzati"
t.font = Font(name="Calibri", bold=True, size=16, color=C_WHITE)
t.fill = fill(C_DARK_BLUE)
t.alignment = align("center", "center", False)
ws.row_dimensions[1].height = 36

# Column headers
col_headers = ["ID", "Modulo", "File di Test", "Funzione", "Descrizione", "Passi", "Risultato Atteso"]
for col, h in enumerate(col_headers, 1):
    c = ws.cell(row=2, column=col)
    c.value = h
    c.font = Font(name="Calibri", bold=True, size=11, color=C_WHITE)
    c.fill = fill(C_MID_BLUE)
    c.alignment = align("center", "center", False)
    c.border = BORDER_MED_TOP
ws.row_dimensions[2].height = 22
ws.freeze_panes = "A3"

# Data rows
prev_module = None
mod_row_counter = {}

for test in TESTS:
    tid, modulo, fname, func, desc, passi, risultato = test
    row_num = tid + 2  # row 3 = test 1

    bg_h, bg_even = MODULE_STYLE[modulo]
    mod_row_counter[modulo] = mod_row_counter.get(modulo, 0) + 1
    use_even = mod_row_counter[modulo] % 2 == 1

    row_bg = bg_even if use_even else C_WHITE

    values = [tid, modulo, fname, func, desc, passi, risultato]
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row_num, column=col)
        c.value = val
        c.fill = fill(row_bg)
        c.border = BORDER_THIN
        if col == 1:  # ID
            c.font = Font(name="Calibri", bold=True, size=10, color=bg_h)
            c.alignment = align("center", "center", False)
        elif col == 2:  # Modulo
            c.font = Font(name="Calibri", bold=True, size=10, color=bg_h)
            c.alignment = align("center", "center", False)
        elif col == 3:  # File
            c.font = Font(name="Courier New", size=9, color="333333")
            c.alignment = align("left", "center", False)
        elif col == 4:  # Funzione
            c.font = Font(name="Courier New", bold=True, size=9, color="1A1A2E")
            c.alignment = align("left", "center", False)
        elif col == 5:  # Descrizione
            c.font = Font(name="Calibri", size=10, color="1A1A1A")
            c.alignment = align("left", "top", True)
        elif col == 6:  # Passi
            c.font = Font(name="Calibri", size=9.5, color="1A1A1A")
            c.alignment = align("left", "top", True)
        elif col == 7:  # Risultato
            c.font = Font(name="Calibri", italic=True, size=9.5, color="1A3A1A")
            c.alignment = align("left", "top", True)

    # Row height based on estimated visual lines (considers text wrap per column)
    def vlines(text, col_w):
        if not text:
            return 1
        return sum(max(1, (len(line) // col_w) + 1) for line in str(text).split("\n"))
    max_vl = max(
        vlines(passi, col_widths[5]),       # Passi (width 50)
        vlines(desc, col_widths[4]),         # Descrizione (width 60)
        vlines(risultato, col_widths[6]),    # Risultato Atteso (width 60)
    )
    ws.row_dimensions[row_num].height = max(40, min(max_vl * 14, 409))


# ── ONE SHEET PER MODULE ───────────────────────────────────────────────────────
for mod_name in ["Login", "Messaggi", "Calendario", "Contatti", "Impostazioni",
                  "Archivio", "Conservazione", "Fatture", "Accessi"]:
    bg_h, bg_even = MODULE_STYLE[mod_name]
    mod_tests = [t for t in TESTS if t[1] == mod_name]

    ws_m = wb.create_sheet(title=mod_name)
    ws_m.sheet_view.showGridLines = False

    col_widths_m = [5, 32, 38, 58, 52, 58]
    for i, w in enumerate(col_widths_m, 1):
        ws_m.column_dimensions[get_column_letter(i)].width = w

    # Title
    ws_m.merge_cells("A1:F1")
    title_m = ws_m["A1"]
    title_m.value = f"MODULO: {mod_name.upper()}  –  {len(mod_tests)} Casi di Test"
    title_m.font = Font(name="Calibri", bold=True, size=14, color=C_WHITE)
    title_m.fill = fill(bg_h)
    title_m.alignment = align("center", "center", False)
    ws_m.row_dimensions[1].height = 32

    # Headers
    col_h_m = ["ID", "File / Funzione", "Descrizione", "Passi", "Risultato Atteso", "Note"]
    for col, h in enumerate(col_h_m, 1):
        c = ws_m.cell(row=2, column=col)
        c.value = h
        c.font = Font(name="Calibri", bold=True, size=10, color=C_WHITE)
        c.fill = fill(bg_h)
        c.alignment = align("center", "center", False)
        c.border = BORDER_THIN
    ws_m.row_dimensions[2].height = 20
    ws_m.freeze_panes = "A3"

    for i, test in enumerate(mod_tests):
        tid, modulo, fname, func, desc, passi, risultato = test
        row_num = i + 3
        row_bg = bg_even if i % 2 == 0 else C_WHITE

        file_func = f"{fname}\n{func}"
        values = [tid, file_func, desc, passi, risultato, ""]
        for col, val in enumerate(values, 1):
            c = ws_m.cell(row=row_num, column=col)
            c.value = val
            c.fill = fill(row_bg)
            c.border = BORDER_THIN
            if col == 1:
                c.font = Font(name="Calibri", bold=True, size=11, color=bg_h)
                c.alignment = align("center", "center", False)
            elif col == 2:
                c.font = Font(name="Courier New", size=9, color="222222")
                c.alignment = align("left", "top", True)
            elif col == 3:
                c.font = Font(name="Calibri", size=10)
                c.alignment = align("left", "top", True)
            elif col == 4:
                c.font = Font(name="Calibri", size=9.5)
                c.alignment = align("left", "top", True)
            elif col == 5:
                c.font = Font(name="Calibri", italic=True, size=9.5, color="1A3A1A")
                c.alignment = align("left", "top", True)
            else:
                c.font = Font(name="Calibri", size=9, color="808080")
                c.alignment = align("left", "top", True)

        def vlines_m(text, col_w):
            if not text:
                return 1
            return sum(max(1, (len(line) // col_w) + 1) for line in str(text).split("\n"))
        max_vl_m = max(
            vlines_m(passi, col_widths_m[3]),       # Passi (width 58)
            vlines_m(desc, col_widths_m[2]),         # Descrizione (width 38)
            vlines_m(risultato, col_widths_m[4]),    # Risultato Atteso (width 52)
            vlines_m(file_func, col_widths_m[1]),    # File/Funzione (width 32)
        )
        ws_m.row_dimensions[row_num].height = max(45, min(max_vl_m * 14, 409))


# ── SAVE ──────────────────────────────────────────────────────────────────────
wb.save(OUTPUT_PATH)
print(f"OK  File Excel creato: {OUTPUT_PATH}")
print(f"    Fogli: {[s.title for s in wb.worksheets]}")
print(f"    Totale test documentati: {len(TESTS)}")
