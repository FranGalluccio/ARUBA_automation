"""
Genera il file Excel con la documentazione completa della suite di test Mobile (PEC Aruba).
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

OUTPUT_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "suite_test_pec_mobile.xlsx")

# ── PALETTE ───────────────────────────────────────────────────────────────────
C_DARK_BLUE   = "1F3864"
C_MID_BLUE    = "2E75B6"
C_LIGHT_BLUE  = "D6E4F0"
C_GREEN       = "375623"
C_LIGHT_GREEN = "E2EFDA"
C_PURPLE      = "4B0082"
C_LIGHT_PUR   = "EAE0F5"
C_TEAL        = "1D5F5F"
C_LIGHT_TEA   = "D9EEE8"
C_ORANGE      = "843C0C"
C_LIGHT_ORA   = "FCE4D6"
C_WHITE       = "FFFFFF"
C_LIGHT_GRAY  = "F2F2F2"
C_GOLD        = "FFD700"

THIN  = Side(style="thin",   color="BFBFBF")
MED   = Side(style="medium", color="595959")
BORDER_THIN    = Border(left=THIN, right=THIN, top=THIN,  bottom=THIN)
BORDER_MED_TOP = Border(left=THIN, right=THIN, top=MED,   bottom=THIN)


def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def align(h="left", v="top", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


# ── TEST DATA ─────────────────────────────────────────────────────────────────
# Columns: (id, modulo, file, funzione, descrizione, passi, risultato_atteso)

TESTS = [

    # ══════════════════════════════════════════════════════════════════════════
    # LOGIN
    # ══════════════════════════════════════════════════════════════════════════
    (1, "Login", "test_login_mobile.py", "test_login_mobile",
     "Verifica che il login su mobile carichi l'inbox e la UI mobile sia presente "
     "(bottom navigation bar e hamburger).",
     "1. Accedere con credenziali valide su viewport mobile.\n"
     "2. Attendere il caricamento della webmail.\n"
     "3. Verificare che l'URL contenga 'INBOX', 'messages' o 'mail'.\n"
     "4. Verificare la visibilità della bottom navigation bar.\n"
     "5. Verificare la presenza del bottone hamburger (≡) nell'header.\n"
     "6. Verificare la presenza di un'intestazione di sezione (h1 o equivalente).\n"
     "7. Scattare screenshot.",
     "L'URL è relativo all'inbox. La bottom navigation bar e il bottone hamburger sono visibili, confermando che la UI mobile è caricata correttamente."),

    # ══════════════════════════════════════════════════════════════════════════
    # CALENDARIO
    # ══════════════════════════════════════════════════════════════════════════
    (2, "Calendario", "test_calendario_mobile.py", "test_navigazione_calendario_mobile",
     "Verifica la navigazione al Calendario via bottom tab e il corretto caricamento della sezione.",
     "1. Accedere con credenziali valide su viewport mobile.\n"
     "2. Cliccare il tab 'Calendario' nella bottom navigation bar.\n"
     "3. Attendere il caricamento della sezione.\n"
     "4. Verificare che l'URL contenga 'calendar'/'calendario' oppure che l'heading 'Calendario' sia visibile.\n"
     "5. Verificare la presenza di elementi calendario (griglia, controlli, pulsante 'Nuovo evento').\n"
     "6. Scattare screenshot.",
     "La sezione Calendario è raggiunta tramite bottom tab. L'URL o l'heading confermano il caricamento. Almeno un elemento UI del calendario è visibile."),

    (3, "Calendario", "test_calendario_mobile.py", "test_crea_elimina_evento_mobile",
     "Crea un nuovo evento dal calendario mobile tramite FAB e verifica l'assenza di errori, poi lo elimina.",
     "1. Accedere con credenziali valide su viewport mobile.\n"
     "2. Navigare al Calendario via bottom tab.\n"
     "3. Scattare screenshot pre-creazione.\n"
     "4. Cliccare il FAB '+' per aprire il form 'Nuovo evento'.\n"
     "5. Scattare screenshot del form.\n"
     "6. Inserire un titolo univoco (con timestamp).\n"
     "7. Cliccare 'Salva'.\n"
     "8. Scattare screenshot post-salvataggio.\n"
     "9. Verificare che NON appaia un toast di errore.\n"
     "10. Cleanup: eliminare l'evento dalla vista Eventi.\n"
     "11. Scattare screenshot finale.",
     "Il form 'Nuovo evento' si apre correttamente. Dopo il salvataggio nessun toast di errore appare. Il cleanup elimina l'evento dalla vista Events."),

    # ══════════════════════════════════════════════════════════════════════════
    # CONTATTI
    # ══════════════════════════════════════════════════════════════════════════
    (4, "Contatti", "test_contatti_mobile.py", "test_navigazione_contatti_mobile",
     "Verifica la navigazione alla sezione Contatti via bottom tab e il corretto caricamento.",
     "1. Accedere con credenziali valide su viewport mobile.\n"
     "2. Cliccare il tab 'Contatti' nella bottom navigation bar.\n"
     "3. Attendere il caricamento della sezione.\n"
     "4. Verificare che l'URL contenga 'contacts'/'contatti' oppure che l'heading 'Contatti' sia visibile.\n"
     "5. Verificare la presenza di un campo ricerca o lista contatti.\n"
     "6. Scattare screenshot.",
     "La sezione Contatti è raggiunta tramite bottom tab. L'URL o l'heading confermano il caricamento. Il campo ricerca o la lista contatti è presente."),

    (5, "Contatti", "test_contatti_mobile.py", "test_crea_cerca_elimina_contatto_mobile",
     "Crea un contatto su mobile, lo cerca tramite barra ricerca, verifica che esista, poi lo elimina.",
     "1. Accedere con credenziali valide su viewport mobile.\n"
     "2. Navigare ai Contatti via bottom tab.\n"
     "3. Cliccare il FAB '+' per aprire il form 'Nuovo contatto'.\n"
     "4. Cliccare 'Procedi' se appare il dialog tipo contatto.\n"
     "5. Compilare nome, cognome ed email univoci (con timestamp).\n"
     "6. Cliccare 'Salva'.\n"
     "7. Aprire la barra di ricerca e cercare il nome del contatto.\n"
     "8. Verificare che il contatto sia trovato (via testo o frame-record-desktop).\n"
     "9. Scattare screenshot del risultato.\n"
     "10. Cleanup: selezionare e eliminare il contatto.\n"
     "11. Scattare screenshot finale.",
     "Il contatto è creato, trovato tramite ricerca e visibile nella lista. Il cleanup elimina il contatto correttamente."),

    # ══════════════════════════════════════════════════════════════════════════
    # MESSAGGI
    # ══════════════════════════════════════════════════════════════════════════
    (6, "Messaggi", "test_messaggi_mobile.py", "test_inbox_navigation_mobile",
     "Verifica la navigazione inbox mobile: lista messaggi, apertura drawer hamburger, "
     "switch a Inviati, ritorno all'inbox via bottom tab.",
     "1. Accedere con credenziali valide su viewport mobile.\n"
     "2. Verificare la presenza del contenitore messaggi.\n"
     "3. Scattare screenshot della inbox.\n"
     "4. Aprire il drawer tramite il bottone hamburger.\n"
     "5. Scattare screenshot del drawer aperto.\n"
     "6. Verificare la presenza delle voci 'In arrivo' e 'Inviati' nel drawer.\n"
     "7. Cliccare 'Inviati' e verificare navigazione.\n"
     "8. Tornare all'inbox via bottom tab 'Messaggi'.\n"
     "9. Verificare che l'URL sia relativo all'inbox (INBOX o messages).\n"
     "10. Scattare screenshot finale.",
     "Il drawer mostra le voci 'In arrivo' e 'Inviati'. La navigazione a Inviati funziona. Il ritorno all'inbox via bottom tab è confermato dall'URL."),

    (7, "Messaggi", "test_messaggi_mobile.py", "test_apri_messaggio_mobile",
     "Apre il primo messaggio in inbox su mobile e verifica il caricamento del dettaglio.",
     "1. Accedere con credenziali valide su viewport mobile.\n"
     "2. Attendere il caricamento dell'inbox.\n"
     "3. Scattare screenshot della lista messaggi.\n"
     "4. Tentare apertura tramite bottone accessibilità 'Passa all'ultimo messaggio ricevuto'.\n"
     "5. Fallback: click alle coordinate centro-alto del viewport.\n"
     "6. Attendere il caricamento del dettaglio.\n"
     "7. Scattare screenshot del dettaglio.\n"
     "8. Verificare che l'URL contenga 'webmail'.\n"
     "9. Scattare screenshot finale.",
     "Il messaggio viene aperto (o tentato l'apertura). L'URL rimane nell'ambito della webmail, confermando che non si è verificato un crash o redirect inatteso."),

    (8, "Messaggi", "test_messaggi_mobile.py", "test_componi_invia_mobile",
     "Compone e invia un messaggio tramite FAB mobile e verifica il toast di conferma.",
     "1. Accedere con credenziali valide su viewport mobile.\n"
     "2. Aprire il compose tramite FAB (open_compose).\n"
     "3. Verificare che la finestra di composizione si sia aperta.\n"
     "4. Compilare destinatario, oggetto e corpo.\n"
     "5. Scattare screenshot del form compilato.\n"
     "6. Cliccare 'Invia'.\n"
     "7. Attendere 3 secondi.\n"
     "8. Verificare la presenza del toast di conferma (aru-toast__message).\n"
     "9. Scattare screenshot finale.",
     "La finestra di composizione si apre correttamente. Dopo l'invio appare il toast di conferma, confermando che il messaggio è stato inviato."),

    # ══════════════════════════════════════════════════════════════════════════
    # IMPOSTAZIONI
    # ══════════════════════════════════════════════════════════════════════════
    (9, "Impostazioni", "test_impostazioni_mobile.py", "test_navigazione_impostazioni_mobile",
     "Naviga alle Impostazioni su mobile (tre metodi: bottom tab, hamburger, URL diretto) "
     "e verifica che la pagina carichi correttamente.",
     "1. Accedere con credenziali valide su viewport mobile.\n"
     "2. Tentare navigazione via bottom tab 'Altro' → Impostazioni.\n"
     "3. Se fallisce: tentare via hamburger drawer → Impostazioni.\n"
     "4. Se fallisce: navigare direttamente all'URL /new/settings/home.\n"
     "5. Scattare screenshot della pagina impostazioni.\n"
     "6. Attendere il caricamento Angular.\n"
     "7. Verificare settings_reached (URL o impostazioni nel titolo).\n"
     "8. Verificare il contenuto (h1 Impostazioni, sezioni Casella PEC, ecc.).\n"
     "9. Scattare screenshot finale.",
     "Le Impostazioni sono raggiunte tramite almeno uno dei tre metodi. La pagina mostra il contenuto atteso (heading o sezioni configurazione)."),

    (10, "Impostazioni", "test_impostazioni_mobile.py", "test_accesso_sezione_impostazioni_mobile",
     "Verifica che le sezioni principali delle impostazioni siano accessibili su mobile: "
     "Casella PEC, Messaggi e scrittura, Account e sicurezza.",
     "1. Accedere con credenziali valide su viewport mobile.\n"
     "2. Navigare direttamente a /new/settings/home.\n"
     "3. Attendere il caricamento.\n"
     "4. Scattare screenshot delle sezioni presenti.\n"
     "5. Verificare la presenza di almeno 2 sezioni tra: Casella PEC, Messaggi e scrittura, Account e sicurezza, Aspetto.\n"
     "6. Cliccare 'Casella PEC' per verificare la navigazione interna.\n"
     "7. Scattare screenshot.\n"
     "8. Verificare che almeno 2 sezioni su 4 siano trovate.\n"
     "9. Scattare screenshot finale.",
     "La pagina Impostazioni mostra almeno 2 sezioni principali. La navigazione interna a 'Casella PEC' funziona correttamente su mobile."),
]

# ── MODULE CONFIG (header color, row-even color) ──────────────────────────────
MODULE_STYLE = {
    "Login":        (C_MID_BLUE, C_LIGHT_BLUE),
    "Calendario":   (C_PURPLE,   C_LIGHT_PUR),
    "Contatti":     (C_ORANGE,   C_LIGHT_ORA),
    "Messaggi":     (C_GREEN,    C_LIGHT_GREEN),
    "Impostazioni": (C_TEAL,     C_LIGHT_TEA),
}

MODULE_DESCRIPTIONS = {
    "Login":        "Test di autenticazione mobile: verifica login, bottom nav bar e hamburger.",
    "Calendario":   "Test sul modulo Calendario mobile: navigazione via bottom tab, creazione e cleanup evento tramite FAB.",
    "Contatti":     "Test sulla rubrica Contatti mobile: navigazione via bottom tab, creazione, ricerca e cleanup contatto.",
    "Messaggi":     "Test sulla gestione messaggi mobile: navigazione inbox/drawer, apertura messaggio, composizione e invio via FAB.",
    "Impostazioni": "Test sulle Impostazioni mobile: navigazione multi-metodo e verifica sezioni principali.",
}

# ── WORKBOOK SETUP ────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

# ── SHEET: INDICE ─────────────────────────────────────────────────────────────
ws_idx = wb.active
ws_idx.title = "Indice"
ws_idx.sheet_view.showGridLines = False
ws_idx.column_dimensions["A"].width = 5
ws_idx.column_dimensions["B"].width = 22
ws_idx.column_dimensions["C"].width = 10
ws_idx.column_dimensions["D"].width = 55

ws_idx.merge_cells("A1:D1")
title_cell = ws_idx["A1"]
title_cell.value = "SUITE DI TEST – WEBMAIL PEC ARUBA (MOBILE)"
title_cell.font = Font(name="Calibri", bold=True, size=18, color=C_WHITE)
title_cell.fill = fill(C_DARK_BLUE)
title_cell.alignment = align("center", "center", False)
ws_idx.row_dimensions[1].height = 42

ws_idx.merge_cells("A2:D2")
sub_cell = ws_idx["A2"]
sub_cell.value = "Documentazione completa dei casi di test automatizzati – Viewport Mobile"
sub_cell.font = Font(name="Calibri", italic=True, size=11, color="808080")
sub_cell.fill = fill("EEF4FB")
sub_cell.alignment = align("center", "center", False)
ws_idx.row_dimensions[2].height = 22

ws_idx.row_dimensions[3].height = 10

headers_idx = ["#", "Modulo", "Totale Test", "Descrizione Modulo"]
for col, h in enumerate(headers_idx, 1):
    c = ws_idx.cell(row=4, column=col)
    c.value = h
    c.font = Font(name="Calibri", bold=True, size=11, color=C_WHITE)
    c.fill = fill(C_DARK_BLUE)
    c.alignment = align("center", "center", False)
    c.border = BORDER_THIN
ws_idx.row_dimensions[4].height = 22

from collections import Counter
module_counts = Counter(t[1] for t in TESTS)

for i, (mod, desc) in enumerate(MODULE_DESCRIPTIONS.items()):
    row = 5 + i
    bg_h, bg_even = MODULE_STYLE[mod]
    row_data = [i + 1, mod, module_counts[mod], desc]
    bg = bg_even if i % 2 == 0 else C_LIGHT_GRAY
    for col, val in enumerate(row_data, 1):
        c = ws_idx.cell(row=row, column=col)
        c.value = val
        c.font = Font(name="Calibri", size=10, color="1A1A1A", bold=(col == 2))
        c.fill = fill(bg)
        c.alignment = align("center" if col in (1, 3) else "left", "center", col == 4)
        c.border = BORDER_THIN
    ws_idx.row_dimensions[row].height = 28

total_row = 4 + len(MODULE_DESCRIPTIONS) + 1
ws_idx.merge_cells(f"A{total_row}:B{total_row}")
tot1 = ws_idx[f"A{total_row}"]
tot1.value = "TOTALE CASI DI TEST"
tot1.font = Font(name="Calibri", bold=True, size=11, color=C_WHITE)
tot1.fill = fill(C_DARK_BLUE)
tot1.alignment = align("center", "center", False)
tot1.border = BORDER_THIN

tot2 = ws_idx[f"C{total_row}"]
tot2.value = len(TESTS)
tot2.font = Font(name="Calibri", bold=True, size=12, color=C_GOLD)
tot2.fill = fill(C_DARK_BLUE)
tot2.alignment = align("center", "center", False)
tot2.border = BORDER_THIN

ws_idx[f"D{total_row}"].fill = fill(C_DARK_BLUE)
ws_idx[f"D{total_row}"].border = BORDER_THIN


# ── SHEET: SUITE COMPLETA ─────────────────────────────────────────────────────
ws = wb.create_sheet(title="Suite Completa")
ws.sheet_view.showGridLines = False

col_widths = [6, 14, 38, 44, 62, 52, 62]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.merge_cells("A1:G1")
t = ws["A1"]
t.value = "SUITE DI TEST – WEBMAIL PEC ARUBA (MOBILE)  |  Casi di Test Automatizzati"
t.font = Font(name="Calibri", bold=True, size=16, color=C_WHITE)
t.fill = fill(C_DARK_BLUE)
t.alignment = align("center", "center", False)
ws.row_dimensions[1].height = 36

col_headers = ["ID", "Modulo", "File di Test", "Funzione", "Descrizione", "Passi", "Risultato Atteso"]
for col, h in enumerate(col_headers, 1):
    c = ws.cell(row=2, column=col)
    c.value = h
    c.font = Font(name="Calibri", bold=True, size=11, color=C_WHITE)
    c.fill = fill(C_MID_BLUE)
    c.alignment = align("center", "center", False)
    c.border = BORDER_MED_TOP
ws.row_dimensions[2].height = 22
ws.freeze_panes = "A3"

mod_row_counter = {}
for test in TESTS:
    tid, modulo, fname, func, desc, passi, risultato = test
    row_num = tid + 2

    bg_h, bg_even = MODULE_STYLE[modulo]
    mod_row_counter[modulo] = mod_row_counter.get(modulo, 0) + 1
    use_even = mod_row_counter[modulo] % 2 == 1
    row_bg = bg_even if use_even else C_WHITE

    values = [tid, modulo, fname, func, desc, passi, risultato]
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row_num, column=col)
        c.value = val
        c.fill = fill(row_bg)
        c.border = BORDER_THIN
        if col == 1:
            c.font = Font(name="Calibri", bold=True, size=10, color=bg_h)
            c.alignment = align("center", "center", False)
        elif col == 2:
            c.font = Font(name="Calibri", bold=True, size=10, color=bg_h)
            c.alignment = align("center", "center", False)
        elif col == 3:
            c.font = Font(name="Courier New", size=9, color="333333")
            c.alignment = align("left", "center", False)
        elif col == 4:
            c.font = Font(name="Courier New", bold=True, size=9, color="1A1A2E")
            c.alignment = align("left", "center", False)
        elif col == 5:
            c.font = Font(name="Calibri", size=10, color="1A1A1A")
            c.alignment = align("left", "top", True)
        elif col == 6:
            c.font = Font(name="Calibri", size=9.5, color="1A1A1A")
            c.alignment = align("left", "top", True)
        elif col == 7:
            c.font = Font(name="Calibri", italic=True, size=9.5, color="1A3A1A")
            c.alignment = align("left", "top", True)

    step_count = passi.count("\n") + 1
    ws.row_dimensions[row_num].height = max(40, step_count * 14)


# ── ONE SHEET PER MODULE ───────────────────────────────────────────────────────
for mod_name in ["Login", "Calendario", "Contatti", "Messaggi", "Impostazioni"]:
    bg_h, bg_even = MODULE_STYLE[mod_name]
    mod_tests = [t for t in TESTS if t[1] == mod_name]

    ws_m = wb.create_sheet(title=mod_name)
    ws_m.sheet_view.showGridLines = False

    col_widths_m = [5, 36, 40, 60, 54, 60]
    for i, w in enumerate(col_widths_m, 1):
        ws_m.column_dimensions[get_column_letter(i)].width = w

    ws_m.merge_cells("A1:F1")
    title_m = ws_m["A1"]
    title_m.value = f"MODULO: {mod_name.upper()}  –  {len(mod_tests)} Casi di Test"
    title_m.font = Font(name="Calibri", bold=True, size=14, color=C_WHITE)
    title_m.fill = fill(bg_h)
    title_m.alignment = align("center", "center", False)
    ws_m.row_dimensions[1].height = 32

    col_h_m = ["ID", "File / Funzione", "Descrizione", "Passi", "Risultato Atteso", "Note"]
    for col, h in enumerate(col_h_m, 1):
        c = ws_m.cell(row=2, column=col)
        c.value = h
        c.font = Font(name="Calibri", bold=True, size=10, color=C_WHITE)
        c.fill = fill(bg_h)
        c.alignment = align("center", "center", False)
        c.border = BORDER_THIN
    ws_m.row_dimensions[2].height = 20
    ws_m.freeze_panes = "A3"

    for i, test in enumerate(mod_tests):
        tid, modulo, fname, func, desc, passi, risultato = test
        row_num = i + 3
        row_bg = bg_even if i % 2 == 0 else C_WHITE

        file_func = f"{fname}\n{func}"
        values = [tid, file_func, desc, passi, risultato, ""]
        for col, val in enumerate(values, 1):
            c = ws_m.cell(row=row_num, column=col)
            c.value = val
            c.fill = fill(row_bg)
            c.border = BORDER_THIN
            if col == 1:
                c.font = Font(name="Calibri", bold=True, size=11, color=bg_h)
                c.alignment = align("center", "center", False)
            elif col == 2:
                c.font = Font(name="Courier New", size=9, color="222222")
                c.alignment = align("left", "top", True)
            elif col == 3:
                c.font = Font(name="Calibri", size=10)
                c.alignment = align("left", "top", True)
            elif col == 4:
                c.font = Font(name="Calibri", size=9.5)
                c.alignment = align("left", "top", True)
            elif col == 5:
                c.font = Font(name="Calibri", italic=True, size=9.5, color="1A3A1A")
                c.alignment = align("left", "top", True)
            else:
                c.font = Font(name="Calibri", size=9, color="808080")
                c.alignment = align("left", "top", True)

        step_count = passi.count("\n") + 1
        ws_m.row_dimensions[row_num].height = max(45, step_count * 14)


# ── SAVE ──────────────────────────────────────────────────────────────────────
wb.save(OUTPUT_PATH)
print(f"OK  File Excel creato: {OUTPUT_PATH}")
print(f"    Fogli: {[s.title for s in wb.worksheets]}")
print(f"    Totale test documentati: {len(TESTS)}")
