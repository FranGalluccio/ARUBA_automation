"""
Genera il file Excel con la documentazione completa della suite di test PEL Aruba.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

OUTPUT_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "suite_test_pel.xlsx")

# ── PALETTE ───────────────────────────────────────────────────────────────────
C_DARK_BLUE   = "1F3864"
C_MID_BLUE    = "2E75B6"
C_LIGHT_BLUE  = "D6E4F0"
C_PURPLE      = "4B0082"
C_LIGHT_PUR   = "EAE0F5"
C_TEAL        = "1D5F5F"
C_LIGHT_TEA   = "D9EEE8"
C_WHITE       = "FFFFFF"
C_LIGHT_GRAY  = "F2F2F2"
C_GOLD        = "FFD700"

THIN  = Side(style="thin",   color="BFBFBF")
MED   = Side(style="medium", color="595959")
BORDER_THIN    = Border(left=THIN, right=THIN, top=THIN,  bottom=THIN)
BORDER_MED_TOP = Border(left=THIN, right=THIN, top=MED,   bottom=THIN)


def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def font(bold=False, color=C_DARK_BLUE, size=10, italic=False):
    return Font(name="Calibri", bold=bold, color=color, size=size, italic=italic)


def align(h="left", v="top", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


# ── TEST DATA ─────────────────────────────────────────────────────────────────
# Columns: (id, modulo, file, funzione, descrizione, passi, risultato_atteso)

TESTS = [

    # ══════════════════════════════════════════════════════════════════════════
    # LOGIN
    # ══════════════════════════════════════════════════════════════════════════
    (1, "Login", "test_login_01.py", "test_logout",
     "Verifica che il logout dalla casella PEL funzioni correttamente.",
     "1. Accedere con credenziali valide.\n"
     "2. Attendere il caricamento della webmail PEL.\n"
     "3. Cliccare il pulsante dell'account (username) in alto a destra.\n"
     "4. Cliccare il pulsante 'Esci' dal menu dropdown.\n"
     "5. Attendere il reindirizzamento alla pagina di login.\n"
     "6. Verificare che il campo username sia visibile.\n"
     "7. Scattare screenshot della pagina di login.",
     "L'URL non contiene più 'INBOX'. Il campo di input username è visibile nella pagina di login."),

    (2, "Login", "test_login_02.py", "test_login_credenziali_errate",
     "Verifica che il login con credenziali non valide mostri un messaggio di errore.",
     "1. Aprire la pagina di login della webmail PEL.\n"
     "2. Inserire un username inesistente nel campo email.\n"
     "3. Inserire una password errata nel campo password.\n"
     "4. Cliccare il pulsante Login.\n"
     "5. Attendere 2 secondi.\n"
     "6. Verificare che l'URL non contenga 'INBOX'.\n"
     "7. Verificare la visibilità del form di login.\n"
     "8. Scattare screenshot.",
     "L'URL rimane sulla pagina di login (no 'INBOX'). Il form di login è ancora visibile, confermando che l'accesso con credenziali errate non è avvenuto."),

    # ══════════════════════════════════════════════════════════════════════════
    # CALENDARIO
    # ══════════════════════════════════════════════════════════════════════════
    (3, "Calendario", "test_calendario_01.py", "test_creazione_evento_ricorrente",
     "Verifica la creazione di un evento ricorrente con regola di ricorrenza personalizzata.",
     "1. Accedere con credenziali valide.\n"
     "2. Navigare nella sezione Calendario.\n"
     "3. Cliccare 'Nuovo evento'.\n"
     "4. Inserire un titolo univoco (con timestamp).\n"
     "5. Aprire il combobox 'Non ripetere' e selezionare 'Personalizza...'.\n"
     "6. Selezionare il radio button 'Dopo' nel dialog di personalizzazione.\n"
     "7. Salvare la regola di ricorrenza.\n"
     "8. Salvare l'evento.\n"
     "9. Verificare la presenza dell'evento ricorrente nel calendario.\n"
     "10. Eliminare l'evento (singolo o tutti).\n"
     "11. Scattare screenshot.",
     "L'evento ricorrente è creato con la regola personalizzata e visibile nel calendario. Il cleanup rimuove correttamente l'evento dalla griglia."),

    (4, "Calendario", "test_calendario_02.py", "test_creazione_modifica_evento",
     "Verifica la creazione di un evento e la successiva modifica del titolo.",
     "1. Accedere con credenziali valide.\n"
     "2. Navigare nella sezione Calendario.\n"
     "3. Cliccare 'Nuovo evento'.\n"
     "4. Inserire un titolo base univoco.\n"
     "5. Salvare l'evento.\n"
     "6. Aprire l'evento dalla griglia (.fc-event).\n"
     "7. Cliccare 'Modifica'.\n"
     "8. Modificare il titolo con uno nuovo univoco.\n"
     "9. Salvare la modifica.\n"
     "10. Verificare che il nuovo titolo sia visibile nel calendario.\n"
     "11. Eliminare l'evento.\n"
     "12. Scattare screenshot.",
     "L'evento è creato con il titolo base. Dopo la modifica, il calendario mostra il nuovo titolo. Il cleanup elimina correttamente l'evento."),

    (5, "Calendario", "test_calendario_03.py", "test_creazione_invio_evento",
     "Verifica la creazione di un evento con invito a partecipanti e la visualizzazione nel calendario.",
     "1. Accedere con credenziali valide.\n"
     "2. Navigare nella sezione Calendario.\n"
     "3. Cliccare 'Nuovo evento'.\n"
     "4. Inserire un titolo univoco.\n"
     "5. Aggiungere un partecipante come invitato (email username).\n"
     "6. Salvare e inviare l'invito.\n"
     "7. Verificare che l'evento sia visibile nel calendario (.fc-event).\n"
     "8. Eliminare l'evento.\n"
     "9. Scattare screenshot.",
     "L'evento è creato con invito e appare nella griglia del calendario. Il cleanup rimuove l'evento."),

    (6, "Calendario", "test_calendario_04.py", "test_import_export_calendario",
     "Verifica l'importazione di un calendario da file ICS e il tentativo di esportazione.",
     "1. Accedere con credenziali valide.\n"
     "2. Navigare nella sezione Calendario.\n"
     "3. Cliccare il pulsante 'Importa'.\n"
     "4. Impostare il file ICS tramite il file input.\n"
     "5. Confermare l'importazione.\n"
     "6. Tentare l'esportazione del calendario (menu contestuale o pulsante).\n"
     "7. Scattare screenshot.",
     "Il pulsante 'Importa' è accessibile. Il file ICS viene caricato senza errori. L'esportazione (se disponibile) avvia il download di un file ICS."),

    (7, "Calendario", "test_calendario_05.py", "test_crea_calendario_personalizzato",
     "Verifica la creazione di un nuovo calendario con nome e colore, poi lo elimina.",
     "1. Accedere con credenziali valide.\n"
     "2. Navigare nella sezione Calendario.\n"
     "3. Aprire 'Gestione calendario'.\n"
     "4. Cliccare 'Nuovo' per creare un nuovo calendario.\n"
     "5. Inserire un nome univoco (con timestamp).\n"
     "6. Selezionare un colore tra quelli disponibili.\n"
     "7. Salvare il calendario.\n"
     "8. Verificare il toast di conferma.\n"
     "9. Verificare che il calendario sia visibile nella sidebar.\n"
     "10. Eliminare il calendario (click destro → Elimina calendario).\n"
     "11. Scattare screenshot.",
     "Il nuovo calendario è creato e appare nella sidebar. Il toast di conferma è visibile. Il calendario viene eliminato correttamente nel cleanup."),

    (8, "Calendario", "test_calendario_06.py", "test_cambio_vista_calendario",
     "Verifica il cambio tra le diverse viste del calendario: Giorno, Mese, Eventi, Settimana.",
     "1. Accedere con credenziali valide.\n"
     "2. Navigare nella sezione Calendario.\n"
     "3. Cliccare il pulsante vista 'Giorno' ([title='Giorno']).\n"
     "4. Scattare screenshot della vista Giorno.\n"
     "5. Cliccare il pulsante vista 'Mese' ([title='Mese']).\n"
     "6. Scattare screenshot della vista Mese.\n"
     "7. Cliccare il pulsante vista 'Eventi' ([title='Eventi']).\n"
     "8. Scattare screenshot della vista Eventi.\n"
     "9. Cliccare il pulsante vista 'Settimana' ([title='Settimana']).\n"
     "10. Scattare screenshot finale.",
     "Il calendario cambia correttamente tra le viste Giorno, Mese, Eventi e Settimana senza errori. Per ogni cambio lo screenshot mostra il layout corretto."),

    (9, "Calendario", "test_calendario_07.py", "test_navigazione_periodi_calendario",
     "Verifica la navigazione del calendario tramite mini-calendar e pulsante 'Oggi'.",
     "1. Accedere con credenziali valide.\n"
     "2. Navigare nella sezione Calendario.\n"
     "3. Leggere il primo giorno della settimana corrente (.fc-col-header-cell).\n"
     "4. Cliccare 'Next' nella mini-calendar per avanzare di un mese.\n"
     "5. Cliccare il primo giorno disponibile del mese successivo.\n"
     "6. Verificare che la main view si sia aggiornata (primo giorno diverso).\n"
     "7. Cliccare il pulsante 'Oggi'.\n"
     "8. Verificare il ritorno alla settimana corrente (primo giorno uguale all'iniziale).\n"
     "9. Scattare screenshot.",
     "La navigazione tramite mini-calendar aggiorna la main view. Il pulsante 'Oggi' riporta correttamente alla settimana corrente."),

    (10, "Calendario", "test_calendario_08.py", "test_evento_giornata_intera",
     "Verifica la creazione di un evento con toggle 'Giornata intera' attivato.",
     "1. Accedere con credenziali valide.\n"
     "2. Navigare nella sezione Calendario.\n"
     "3. Cliccare 'Nuovo evento'.\n"
     "4. Inserire un titolo univoco.\n"
     "5. Attivare il toggle/checkbox 'Giornata intera'.\n"
     "6. Salvare l'evento.\n"
     "7. Verificare la presenza dell'evento nella griglia del calendario.\n"
     "8. Eliminare l'evento.\n"
     "9. Scattare screenshot.",
     "L'evento è creato come evento 'Giornata intera' ed è visibile nel calendario (nella banda all-day o nella griglia). Il cleanup elimina l'evento correttamente."),

    (11, "Calendario", "test_calendario_09.py", "test_evento_con_promemoria",
     "Verifica la configurazione e il salvataggio di un promemoria per un evento del calendario.",
     "1. Accedere con credenziali valide.\n"
     "2. Navigare nella sezione Calendario.\n"
     "3. Cliccare 'Nuovo evento'.\n"
     "4. Inserire un titolo univoco.\n"
     "5. Cliccare 'Aggiungi promemoria'.\n"
     "6. Configurare il tipo di notifica (email) e inserire l'indirizzo.\n"
     "7. Salvare il dialog promemoria.\n"
     "8. Salvare l'evento.\n"
     "9. Verificare che l'evento sia visibile nel calendario (.fc-event).\n"
     "10. Eliminare l'evento.\n"
     "11. Scattare screenshot.",
     "Il promemoria è configurato e l'evento è salvato correttamente. L'evento è visibile nella griglia del calendario. Il cleanup elimina l'evento."),

    (12, "Calendario", "test_calendario_10.py", "test_ricerca_nel_calendario",
     "Verifica la funzione di ricerca eventi nel calendario tramite la barra di ricerca.",
     "1. Accedere con credenziali valide.\n"
     "2. Navigare nella sezione Calendario.\n"
     "3. Creare un evento con titolo univoco (timestamp).\n"
     "4. Salvare l'evento.\n"
     "5. Cercare l'evento tramite la barra 'Cerca nel calendario'.\n"
     "6. Premere Enter per avviare la ricerca.\n"
     "7. Verificare che il testo univoco sia trovato.\n"
     "8. Fallback: se la ricerca restituisce spinner, verificare tramite vista Eventi.\n"
     "9. Eliminare l'evento.\n"
     "10. Scattare screenshot.",
     "L'evento con il titolo univoco è trovato tramite ricerca o visibile nella vista Eventi. Il cleanup elimina l'evento."),

    (13, "Calendario", "test_calendario_11.py", "test_evento_luogo_indirizzo",
     "Verifica la creazione di un evento con tipo luogo 'Indirizzo' e il salvataggio del campo location.",
     "1. Accedere con credenziali valide.\n"
     "2. Navigare nella sezione Calendario.\n"
     "3. Cliccare 'Nuovo evento'.\n"
     "4. Inserire un titolo univoco.\n"
     "5. Inserire un indirizzo fisico nel campo luogo (es. 'Via Roma 1, Milano').\n"
     "6. Salvare l'evento.\n"
     "7. Aprire l'evento dalla griglia e verificare che l'indirizzo sia presente nel popup.\n"
     "8. Eliminare l'evento.\n"
     "9. Scattare screenshot.",
     "Il luogo 'Via Roma 1, Milano' è salvato correttamente e visibile nel popup dell'evento. Il cleanup elimina l'evento."),

    (14, "Calendario", "test_calendario_12.py", "test_evento_luogo_link",
     "Verifica la creazione di un evento con URL meeting come luogo e il salvataggio del campo.",
     "1. Accedere con credenziali valide.\n"
     "2. Navigare nella sezione Calendario.\n"
     "3. Cliccare 'Nuovo evento'.\n"
     "4. Inserire un titolo univoco.\n"
     "5. Inserire un URL meeting nel campo luogo (es. 'https://meet.example.com/test-room').\n"
     "6. Salvare l'evento.\n"
     "7. Aprire l'evento dalla griglia e verificare che l'URL sia presente nel popup.\n"
     "8. Eliminare l'evento.\n"
     "9. Scattare screenshot.",
     "L'URL meeting è salvato correttamente e visibile (parzialmente o intero) nel popup dell'evento. Il cleanup elimina l'evento."),

    (15, "Calendario", "test_calendario_13.py", "test_evento_corpo_rich_text",
     "Verifica la creazione di un evento con corpo in rich text (grassetto + lista) tramite suneditor.",
     "1. Accedere con credenziali valide.\n"
     "2. Navigare nella sezione Calendario.\n"
     "3. Cliccare 'Nuovo evento'.\n"
     "4. Inserire un titolo univoco.\n"
     "5. Espandere a fullscreen per accedere al rich text editor.\n"
     "6. Cliccare nell'area editable del suneditor.\n"
     "7. Attivare il grassetto (toolbar o Ctrl+B) e digitare il testo.\n"
     "8. Aggiungere una voce di lista.\n"
     "9. Salvare l'evento.\n"
     "10. Aprire l'evento e verificare che il testo in grassetto sia visibile nel popup.\n"
     "11. Eliminare l'evento.\n"
     "12. Scattare screenshot.",
     "Il testo in grassetto inserito tramite rich text editor è salvato e visibile nel popup evento. Il cleanup elimina l'evento."),

    (16, "Calendario", "test_calendario_14.py", "test_assegna_calendario_specifico",
     "Verifica la creazione di un evento assegnandolo a un calendario specifico.",
     "1. Accedere con credenziali valide.\n"
     "2. Navigare nella sezione Calendario.\n"
     "3. Cliccare 'Nuovo evento'.\n"
     "4. Inserire un titolo univoco.\n"
     "5. Espandere a fullscreen per accedere al campo 'Calendario assegnato'.\n"
     "6. Aprire il combobox del calendario (default 'Personale').\n"
     "7. Selezionare un'opzione diversa dall'attuale se disponibile.\n"
     "8. Salvare l'evento.\n"
     "9. Verificare che l'evento sia visibile nel calendario.\n"
     "10. Eliminare l'evento.\n"
     "11. Scattare screenshot.",
     "L'evento è creato e assegnato al calendario selezionato. L'evento è visibile nella griglia del calendario. Il cleanup elimina l'evento correttamente."),

    (17, "Calendario", "test_calendario_15.py", "test_disponibilita_occupato",
     "Verifica che un evento con 'Mostrati come occupato' ON salvi correttamente la preferenza.",
     "1. Accedere con credenziali valide.\n"
     "2. Navigare nella sezione Calendario.\n"
     "3. Creare un nuovo evento con orario futuro (+2h).\n"
     "4. Espandere a fullscreen per accedere al toggle 'Mostrati come occupato'.\n"
     "5. Assicurarsi che il toggle sia attivo (ON).\n"
     "6. Salvare l'evento.\n"
     "7. Riaprire l'evento dalla griglia.\n"
     "8. Cliccare 'Modifica'.\n"
     "9. Verificare che il toggle 'Mostrati come occupato' risulti checked.\n"
     "10. Chiudere senza salvare.\n"
     "11. Eliminare l'evento nel finally.\n"
     "12. Scattare screenshot.",
     "Il toggle 'Mostrati come occupato' è checked (ON) alla riapertura dell'evento, confermando che la preferenza è stata salvata correttamente."),

    (18, "Calendario", "test_calendario_16.py", "test_disponibilita_libero",
     "Verifica che un evento con 'Mostrati come occupato' OFF non renda lo slot occupato nella vista pianificazione.",
     "1. Accedere con credenziali valide.\n"
     "2. Creare evento A con 'Mostrati come occupato' OFF (Libero) alle 14:00-15:00.\n"
     "3. Creare evento B alla stessa ora aggiungendo se stesso come invitato.\n"
     "4. Aprire la vista di pianificazione disponibilità (aru-button.planning-button).\n"
     "5. Verificare che lo slot 14:00-15:00 NON risulti 'Occupato' (exact=True).\n"
     "6. Chiudere la vista pianificazione e annullare l'evento B.\n"
     "7. Eliminare entrambi gli eventi nel finally.\n"
     "8. Scattare screenshot.",
     "Lo slot 14:00-15:00 non risulta 'Occupato' nella vista pianificazione, confermando che l'evento con 'Mostrati come occupato' OFF non impatta la disponibilità."),

    # ══════════════════════════════════════════════════════════════════════════
    # IMPOSTAZIONI
    # ══════════════════════════════════════════════════════════════════════════
    (19, "Impostazioni", "test_impostazioni_calendari.py", "test_impostazioni_luogo_predefinito",
     "S1: Imposta un luogo predefinito nelle impostazioni calendario e verifica che venga pre-compilato alla creazione di un nuovo evento.",
     "1. Accedere con credenziali valide.\n"
     "2. Navigare a /settings/calendars/customization tramite URL diretto.\n"
     "3. Scattare screenshot della pagina impostazioni.\n"
     "4. Cercare il campo luogo (input[aria-label='input field'] o placeholder 'Luogo').\n"
     "5. Se il campo non è trovato: verificare che la pagina impostazioni sia accessibile e terminare (PASS).\n"
     "6. Se trovato: inserire 'Sala Riunioni Test' nel campo luogo.\n"
     "7. Salvare le impostazioni.\n"
     "8. Navigare al calendario e creare un nuovo evento.\n"
     "9. Leggere il valore del campo luogo nel form evento.\n"
     "10. Chiudere il form senza salvare.\n"
     "11. Ripristinare il luogo vuoto nelle impostazioni.\n"
     "12. Scattare screenshot.",
     "La pagina impostazioni calendario è accessibile. Se il campo luogo è presente, il valore viene salvato e letto nel form del nuovo evento. L'impostazione viene ripristinata al termine."),

    (20, "Impostazioni", "test_impostazioni_calendari.py", "test_impostazioni_visualizzazione_predefinita",
     "S2: Cambia la visualizzazione predefinita del calendario e verifica che il calendario si apra con quella vista.",
     "1. Accedere con credenziali valide.\n"
     "2. Navigare a /settings/calendars/customization tramite URL diretto.\n"
     "3. Scattare screenshot della pagina impostazioni.\n"
     "4. Leggere la visualizzazione corrente via JS shadow DOM.\n"
     "5. Aprire il combobox 'Visualizzazione predefinita' cliccando tutti i [role='combobox'] via JS.\n"
     "6. Cercare e cliccare l'opzione 'Mese' (via Playwright o JS shadow DOM traversal).\n"
     "7. Salvare le impostazioni.\n"
     "8. Navigare al calendario e verificare che la vista Mese sia attiva.\n"
     "9. Ripristinare la visualizzazione predefinita originale.\n"
     "10. Scattare screenshot.",
     "Se la selezione della vista 'Mese' è riuscita: il calendario si apre in vista Mese. Se il combobox usa shadow DOM non interagibile: la pagina impostazioni è accessibile (PASS con fallback)."),
]

# ── MODULE CONFIG (header color, row-even color) ──────────────────────────────
MODULE_STYLE = {
    "Login":        (C_MID_BLUE, C_LIGHT_BLUE),
    "Calendario":   (C_PURPLE,   C_LIGHT_PUR),
    "Impostazioni": (C_TEAL,     C_LIGHT_TEA),
}

MODULE_DESCRIPTIONS = {
    "Login":        "Test di autenticazione: verifica logout e login con credenziali errate.",
    "Calendario":   "Test sul modulo Calendario PEL: creazione/modifica eventi, ricorrenze, viste, navigazione, giornata intera, promemoria, ricerca, luogo, rich text, calendario assegnato e disponibilità.",
    "Impostazioni": "Test sulle impostazioni del calendario PEL: luogo predefinito (S1) e visualizzazione predefinita (S2).",
}

# ── WORKBOOK SETUP ────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

# ── SHEET: INDICE ─────────────────────────────────────────────────────────────
ws_idx = wb.active
ws_idx.title = "Indice"
ws_idx.sheet_view.showGridLines = False
ws_idx.column_dimensions["A"].width = 5
ws_idx.column_dimensions["B"].width = 22
ws_idx.column_dimensions["C"].width = 10
ws_idx.column_dimensions["D"].width = 55

# Title
ws_idx.merge_cells("A1:D1")
title_cell = ws_idx["A1"]
title_cell.value = "SUITE DI TEST – WEBMAIL PEL ARUBA"
title_cell.font = Font(name="Calibri", bold=True, size=18, color=C_WHITE)
title_cell.fill = fill(C_DARK_BLUE)
title_cell.alignment = align("center", "center", False)
ws_idx.row_dimensions[1].height = 42

ws_idx.merge_cells("A2:D2")
sub_cell = ws_idx["A2"]
sub_cell.value = "Documentazione completa dei casi di test automatizzati"
sub_cell.font = Font(name="Calibri", italic=True, size=11, color="808080")
sub_cell.fill = fill("EEF4FB")
sub_cell.alignment = align("center", "center", False)
ws_idx.row_dimensions[2].height = 22

ws_idx.row_dimensions[3].height = 10

# Headers
headers_idx = ["#", "Modulo", "Totale Test", "Descrizione Modulo"]
for col, h in enumerate(headers_idx, 1):
    c = ws_idx.cell(row=4, column=col)
    c.value = h
    c.font = Font(name="Calibri", bold=True, size=11, color=C_WHITE)
    c.fill = fill(C_DARK_BLUE)
    c.alignment = align("center", "center", False)
    c.border = BORDER_THIN
ws_idx.row_dimensions[4].height = 22

from collections import Counter
module_counts = Counter(t[1] for t in TESTS)

for i, (mod, desc) in enumerate(MODULE_DESCRIPTIONS.items()):
    row = 5 + i
    bg_h, bg_even = MODULE_STYLE[mod]
    row_data = [i + 1, mod, module_counts[mod], desc]
    bg = bg_even if i % 2 == 0 else C_LIGHT_GRAY
    for col, val in enumerate(row_data, 1):
        c = ws_idx.cell(row=row, column=col)
        c.value = val
        c.font = Font(name="Calibri", size=10, color="1A1A1A", bold=(col == 2))
        c.fill = fill(bg)
        c.alignment = align("center" if col in (1, 3) else "left", "center", col == 4)
        c.border = BORDER_THIN
    ws_idx.row_dimensions[row].height = 28

# Total row
total_row = 4 + len(MODULE_DESCRIPTIONS) + 1
ws_idx.merge_cells(f"A{total_row}:B{total_row}")
tot1 = ws_idx[f"A{total_row}"]
tot1.value = "TOTALE CASI DI TEST"
tot1.font = Font(name="Calibri", bold=True, size=11, color=C_WHITE)
tot1.fill = fill(C_DARK_BLUE)
tot1.alignment = align("center", "center", False)
tot1.border = BORDER_THIN

tot2 = ws_idx[f"C{total_row}"]
tot2.value = len(TESTS)
tot2.font = Font(name="Calibri", bold=True, size=12, color=C_GOLD)
tot2.fill = fill(C_DARK_BLUE)
tot2.alignment = align("center", "center", False)
tot2.border = BORDER_THIN

ws_idx[f"D{total_row}"].fill = fill(C_DARK_BLUE)
ws_idx[f"D{total_row}"].border = BORDER_THIN


# ── SHEET: SUITE COMPLETA ─────────────────────────────────────────────────────
ws = wb.create_sheet(title="Suite Completa")
ws.sheet_view.showGridLines = False

col_widths = [6, 14, 38, 44, 62, 52, 62]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.merge_cells("A1:G1")
t = ws["A1"]
t.value = "SUITE DI TEST – WEBMAIL PEL ARUBA  |  Casi di Test Automatizzati"
t.font = Font(name="Calibri", bold=True, size=16, color=C_WHITE)
t.fill = fill(C_DARK_BLUE)
t.alignment = align("center", "center", False)
ws.row_dimensions[1].height = 36

col_headers = ["ID", "Modulo", "File di Test", "Funzione", "Descrizione", "Passi", "Risultato Atteso"]
for col, h in enumerate(col_headers, 1):
    c = ws.cell(row=2, column=col)
    c.value = h
    c.font = Font(name="Calibri", bold=True, size=11, color=C_WHITE)
    c.fill = fill(C_MID_BLUE)
    c.alignment = align("center", "center", False)
    c.border = BORDER_MED_TOP
ws.row_dimensions[2].height = 22
ws.freeze_panes = "A3"

mod_row_counter = {}
for test in TESTS:
    tid, modulo, fname, func, desc, passi, risultato = test
    row_num = tid + 2

    bg_h, bg_even = MODULE_STYLE[modulo]
    mod_row_counter[modulo] = mod_row_counter.get(modulo, 0) + 1
    use_even = mod_row_counter[modulo] % 2 == 1
    row_bg = bg_even if use_even else C_WHITE

    values = [tid, modulo, fname, func, desc, passi, risultato]
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row_num, column=col)
        c.value = val
        c.fill = fill(row_bg)
        c.border = BORDER_THIN
        if col == 1:
            c.font = Font(name="Calibri", bold=True, size=10, color=bg_h)
            c.alignment = align("center", "center", False)
        elif col == 2:
            c.font = Font(name="Calibri", bold=True, size=10, color=bg_h)
            c.alignment = align("center", "center", False)
        elif col == 3:
            c.font = Font(name="Courier New", size=9, color="333333")
            c.alignment = align("left", "center", False)
        elif col == 4:
            c.font = Font(name="Courier New", bold=True, size=9, color="1A1A2E")
            c.alignment = align("left", "center", False)
        elif col == 5:
            c.font = Font(name="Calibri", size=10, color="1A1A1A")
            c.alignment = align("left", "top", True)
        elif col == 6:
            c.font = Font(name="Calibri", size=9.5, color="1A1A1A")
            c.alignment = align("left", "top", True)
        elif col == 7:
            c.font = Font(name="Calibri", italic=True, size=9.5, color="1A3A1A")
            c.alignment = align("left", "top", True)

    step_count = passi.count("\n") + 1
    ws.row_dimensions[row_num].height = max(40, step_count * 14)


# ── ONE SHEET PER MODULE ───────────────────────────────────────────────────────
for mod_name in ["Login", "Calendario", "Impostazioni"]:
    bg_h, bg_even = MODULE_STYLE[mod_name]
    mod_tests = [t for t in TESTS if t[1] == mod_name]

    ws_m = wb.create_sheet(title=mod_name)
    ws_m.sheet_view.showGridLines = False

    col_widths_m = [5, 36, 40, 60, 54, 60]
    for i, w in enumerate(col_widths_m, 1):
        ws_m.column_dimensions[get_column_letter(i)].width = w

    ws_m.merge_cells("A1:F1")
    title_m = ws_m["A1"]
    title_m.value = f"MODULO: {mod_name.upper()}  –  {len(mod_tests)} Casi di Test"
    title_m.font = Font(name="Calibri", bold=True, size=14, color=C_WHITE)
    title_m.fill = fill(bg_h)
    title_m.alignment = align("center", "center", False)
    ws_m.row_dimensions[1].height = 32

    col_h_m = ["ID", "File / Funzione", "Descrizione", "Passi", "Risultato Atteso", "Note"]
    for col, h in enumerate(col_h_m, 1):
        c = ws_m.cell(row=2, column=col)
        c.value = h
        c.font = Font(name="Calibri", bold=True, size=10, color=C_WHITE)
        c.fill = fill(bg_h)
        c.alignment = align("center", "center", False)
        c.border = BORDER_THIN
    ws_m.row_dimensions[2].height = 20
    ws_m.freeze_panes = "A3"

    for i, test in enumerate(mod_tests):
        tid, modulo, fname, func, desc, passi, risultato = test
        row_num = i + 3
        row_bg = bg_even if i % 2 == 0 else C_WHITE

        file_func = f"{fname}\n{func}"
        values = [tid, file_func, desc, passi, risultato, ""]
        for col, val in enumerate(values, 1):
            c = ws_m.cell(row=row_num, column=col)
            c.value = val
            c.fill = fill(row_bg)
            c.border = BORDER_THIN
            if col == 1:
                c.font = Font(name="Calibri", bold=True, size=11, color=bg_h)
                c.alignment = align("center", "center", False)
            elif col == 2:
                c.font = Font(name="Courier New", size=9, color="222222")
                c.alignment = align("left", "top", True)
            elif col == 3:
                c.font = Font(name="Calibri", size=10)
                c.alignment = align("left", "top", True)
            elif col == 4:
                c.font = Font(name="Calibri", size=9.5)
                c.alignment = align("left", "top", True)
            elif col == 5:
                c.font = Font(name="Calibri", italic=True, size=9.5, color="1A3A1A")
                c.alignment = align("left", "top", True)
            else:
                c.font = Font(name="Calibri", size=9, color="808080")
                c.alignment = align("left", "top", True)

        step_count = passi.count("\n") + 1
        ws_m.row_dimensions[row_num].height = max(45, step_count * 14)


# ── SAVE ──────────────────────────────────────────────────────────────────────
wb.save(OUTPUT_PATH)
print(f"OK  File Excel creato: {OUTPUT_PATH}")
print(f"    Fogli: {[s.title for s in wb.worksheets]}")
print(f"    Totale test documentati: {len(TESTS)}")
